"""
Build the SMB internal pricing & staffing analysis workbook — v0.2 FINAL.

Aligned to the FINAL commercial agreement
(see context/aideology_eand_final_agreement.md):

  Fixed-fee envelope (10 milestones over 30 weeks):
    Wave 1 — Platform + Customer Agent ............ $1,147,621
    Wave 2 — Sales + Comms Agents .................   $575,000
    Wave 3 — Finance + Ops Agents .................   $575,000
    Wave 4 — People Agent + Platform Hardening ....  $1,146,000
    Total .........................................  $3,443,621

  Revenue share (declining, build-then-transfer):
    Year 1-2:  65% e&  /  35% AIdeology
    Year 3:    72% e&  /  28% AIdeology
    Year 4+:   80% e&  /  20% AIdeology  (licensing)

  Adoption / ARPU / e& gross revenue (per agreement):
    Y1:  24,000 EOY  · ARPU AED 285  · gross AED  36.5M
    Y2:  62,000 EOY  · ARPU AED 350  · gross AED 154.0M
    Y3:  85,000 EOY  · ARPU AED 420  · gross AED 378.0M

  AIdeology cash projection (per agreement):
    Y1: $3.44M fixed + AED 12.8M rev-share = ~$7.1M
    Y2: AED 53.9M rev-share                = ~$15.4M
    Y3: AED 105.8M rev-share               = ~$30.2M
    3-Year Total: ~$52.7M  (14.5x return)

This workbook is the INTERNAL view for AIdeology — it derives our
build cost from first principles and confirms the margin we earn
on each Wave at the agreed fees, then layers the rev-share + the
acquisition/valuation pathway on top.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "Internal Analysis/AIdeology_eand_SMB_Pricing_Analysis_v0.1.xlsx"

# -----------------------------------------------------------------------------
# Style palette
# -----------------------------------------------------------------------------
NAVY  = "FF1F3A5F"
WHITE = "FFFFFFFF"
GREY  = "FF808080"
INK   = "FF000000"
BLUE  = "FF0000FF"
GREEN = "FF008000"
RED   = "FFC8102E"
LIGHT_YELLOW = "FFFFF2CC"
LIGHT_GREY   = "FFF5F8FB"
LIGHT_GREEN  = "FFE2F0D9"
LIGHT_BLUE   = "FFDDEBF7"

thin = Side(border_style="thin", color="FFD0D5DC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

F_TITLE = Font(name="Calibri", size=16, bold=True, color=NAVY)
F_SUB   = Font(name="Calibri", size=11, color=GREY, italic=True)
F_NOTE  = Font(name="Calibri", size=10, color=GREY)
F_H1    = Font(name="Calibri", size=12, bold=True, color="FF2E75B6")
F_HDR   = Font(name="Calibri", size=10, bold=True, color=WHITE)
F_LBL   = Font(name="Calibri", size=10, bold=True, color=INK)
F_TXT   = Font(name="Calibri", size=10, color=INK)
F_INPUT = Font(name="Calibri", size=10, bold=False, color=BLUE)
F_FORMULA = Font(name="Calibri", size=10, color=GREEN)
F_TOT   = Font(name="Calibri", size=10, bold=True, color=INK)
F_BIG   = Font(name="Calibri", size=24, bold=True, color=NAVY)
F_MID   = Font(name="Calibri", size=14, bold=True, color=NAVY)
F_DESC  = Font(name="Calibri", size=11, color="FF444444")
F_VAL_BIG = Font(name="Calibri", size=14, bold=True, color=NAVY)
F_VAL   = Font(name="Calibri", size=11, bold=True, color=INK)
F_LABEL = Font(name="Calibri", size=11, color=GREY)

FILL_HDR    = PatternFill("solid", fgColor=NAVY)
FILL_INPUT  = PatternFill("solid", fgColor=LIGHT_YELLOW)
FILL_DATA   = PatternFill("solid", fgColor=LIGHT_GREY)
FILL_TOTAL  = PatternFill("solid", fgColor=LIGHT_GREEN)
FILL_SUB    = PatternFill("solid", fgColor=LIGHT_BLUE)
FILL_ACCENT = PatternFill("solid", fgColor="FFE8F0FE")
FILL_GREEN_LT = PatternFill("solid", fgColor="FFE8F5E9")
FILL_WARN   = PatternFill("solid", fgColor="FFFFF3E0")
FILL_GOLD   = PatternFill("solid", fgColor="FFFFF8E1")

ALIGN_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right",  vertical="center", wrap_text=True)


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def style_input(c):
    c.font = F_INPUT; c.fill = FILL_INPUT; c.border = BORDER; c.alignment = ALIGN_R

def style_formula(c):
    c.font = F_FORMULA; c.fill = FILL_DATA; c.border = BORDER; c.alignment = ALIGN_R

def style_total(c):
    c.font = F_TOT; c.fill = FILL_TOTAL; c.border = BORDER; c.alignment = ALIGN_R

def style_label(c):
    c.font = F_LBL; c.alignment = ALIGN_L; c.border = BORDER

def style_text(c):
    c.font = F_TXT; c.alignment = ALIGN_L; c.border = BORDER

def set_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def section_banner(ws, row, col_start, col_end, text, fill=FILL_HDR, font_color=WHITE):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = Font(name="Calibri", size=13, bold=True, color=font_color)
    c.fill = fill; c.alignment = ALIGN_L; c.border = BORDER
    for col in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).border = BORDER

def summary_box(ws, row, col, label, formula, fmt="$#,##0", width=2, accent=FILL_ACCENT):
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = F_LABEL; lc.alignment = ALIGN_L; lc.border = BORDER; lc.fill = accent
    vc = ws.cell(row=row + 1, column=col, value=formula)
    vc.font = F_VAL_BIG; vc.alignment = ALIGN_L; vc.border = BORDER
    vc.number_format = fmt; vc.fill = accent
    if width > 1:
        for c in range(col + 1, col + width):
            ws.cell(row=row, column=c).fill = accent
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row + 1, column=c).fill = accent
            ws.cell(row=row + 1, column=c).border = BORDER
# =============================================================================
wb = Workbook()
wb.remove(wb.active)


# =============================================================================
# SHEET 0 — Executive Summary
# =============================================================================
ws = wb.create_sheet("0. Executive Summary")
set_widths(ws, {"A": 3, "B": 30, "C": 22, "D": 22, "E": 22, "F": 22, "G": 22})

ws.merge_cells("B2:G2")
ws["B2"] = "AIdeology × e&  —  SMB Internal Pricing Summary  (FINAL framework)"
ws["B2"].font = F_BIG; ws["B2"].alignment = ALIGN_L

ws.merge_cells("B3:G3")
ws["B3"] = "One page. The deal we are signing. Detail in tabs 1-11."
ws["B3"].font = F_DESC; ws["B3"].alignment = ALIGN_L

# ── A. WHAT ARE WE BUILDING? ──
section_banner(ws, 5, 2, 7, "A.  WHAT ARE WE BUILDING?")
desc_lines = [
    "AIdeology builds an AI orchestration platform for e&. SMBs subscribe to pre-built agents (Customer, Sales, Comms, Finance, Ops, People) on top of e&'s monthly invoice — paid like any other telco add-on.",
    "Wave 1 = base platform + Customer Agent (12 weeks). Waves 2-4 = the other 5 agents in three packages (Sales+Comms, Finance+Ops, People+hardening) over the next 18 weeks.",
    "Build-then-transfer: AIdeology delivers Wave 1-4 then trains the e& AI team. By end of Year 2, e& owns the agent code; AIdeology keeps the platform IP and earns rev-share + future licensing.",
]
for i, line in enumerate(desc_lines):
    ws.merge_cells(start_row=6 + i, start_column=2, end_row=6 + i, end_column=7)
    c = ws.cell(row=6 + i, column=2, value=line); c.font = F_DESC; c.alignment = ALIGN_L

# ── B. WHAT WE CHARGE e& (fixed fees) ──
section_banner(ws, 10, 2, 7, "B.  WHAT WE CHARGE e&  —  FIXED FEES  (10 milestones · 30 weeks)")
summary_box(ws, 11, 2, "Wave 1 (Platform + Customer Agent)", "='7. Pricing Summary'!C9")
summary_box(ws, 11, 4, "Waves 2+3+4 (5 agents + hardening)", "='7. Pricing Summary'!C10+'7. Pricing Summary'!C11+'7. Pricing Summary'!C12")
summary_box(ws, 11, 6, "Total fixed-fee envelope", "='7. Pricing Summary'!C13")

# Margin row
ws.cell(row=14, column=2, value="Our internal cost (Wave 1)").font = F_LABEL
ws.cell(row=14, column=2).alignment = ALIGN_L
v = ws.cell(row=14, column=3, value="='7. Pricing Summary'!D9"); v.font = F_VAL; v.number_format = "$#,##0"
ws.cell(row=14, column=4, value="Our margin (Wave 1)").font = F_LABEL
ws.cell(row=14, column=4).alignment = ALIGN_L
m = ws.cell(row=14, column=5, value="=('7. Pricing Summary'!C9-'7. Pricing Summary'!D9)/'7. Pricing Summary'!C9")
m.font = F_VAL; m.number_format = "0.0%"
ws.cell(row=14, column=6, value="Total programme margin").font = F_LABEL
ws.cell(row=14, column=6).alignment = ALIGN_L
m2 = ws.cell(row=14, column=7, value="=('7. Pricing Summary'!C13-'7. Pricing Summary'!D13)/'7. Pricing Summary'!C13")
m2.font = F_VAL; m2.number_format = "0.0%"

# ── C. RECURRING REVENUE — rev-share ──
section_banner(ws, 16, 2, 7, "C.  RECURRING REVENUE  —  declining rev-share (65/35 → 72/28 → 80/20)")

th_row = 17
for i, h in enumerate(["", "Year 1", "Year 2", "Year 3", "3-Year Total"]):
    c = ws.cell(row=th_row, column=2 + i, value=h)
    c.font = F_LBL; c.fill = FILL_ACCENT; c.alignment = ALIGN_C; c.border = BORDER

# SMB customers EOY
sub_row = th_row + 1
ws.cell(row=sub_row, column=2, value="SMB customers (end of year)").font = F_LBL; ws.cell(row=sub_row, column=2).border = BORDER
for j, src in enumerate(["C", "D", "E"]):
    cell = ws.cell(row=sub_row, column=3 + j, value=f"='8. Revenue Share Model'!{src}7")
    cell.font = F_VAL; cell.number_format = "#,##0"; cell.border = BORDER
ws.cell(row=sub_row, column=6).border = BORDER

# e& gross revenue (AED)
egross_row = sub_row + 1
ws.cell(row=egross_row, column=2, value="e& gross SaaS revenue (AED)").font = F_LBL; ws.cell(row=egross_row, column=2).border = BORDER
for j in range(3):
    col = 3 + j
    cell = ws.cell(row=egross_row, column=col, value=f"='8. Revenue Share Model'!{get_column_letter(col)}11")
    cell.font = F_VAL; cell.number_format = "#,##0,,\"M AED\""; cell.border = BORDER
gt = ws.cell(row=egross_row, column=6, value=f"=SUM(C{egross_row}:E{egross_row})")
gt.font = F_VAL; gt.number_format = "#,##0,,\"M AED\""; gt.border = BORDER

# AIdeology rev-share %
pctrow = egross_row + 1
ws.cell(row=pctrow, column=2, value="Our rev-share % (declining)").font = F_LBL; ws.cell(row=pctrow, column=2).border = BORDER
for j in range(3):
    col = 3 + j
    cell = ws.cell(row=pctrow, column=col, value=f"='8. Revenue Share Model'!{get_column_letter(col)}12")
    cell.font = F_VAL; cell.number_format = "0%"; cell.border = BORDER
ws.cell(row=pctrow, column=6).border = BORDER

# AIdeology rev-share AED
ars_row = pctrow + 1
ws.cell(row=ars_row, column=2, value="Our rev-share revenue (AED)").font = F_LBL
ws.cell(row=ars_row, column=2).border = BORDER; ws.cell(row=ars_row, column=2).fill = FILL_GREEN_LT
for j in range(3):
    col = 3 + j
    cell = ws.cell(row=ars_row, column=col, value=f"='8. Revenue Share Model'!{get_column_letter(col)}13")
    cell.font = F_VAL_BIG; cell.number_format = "#,##0,,\"M AED\""; cell.border = BORDER; cell.fill = FILL_GREEN_LT
gt = ws.cell(row=ars_row, column=6, value=f"=SUM(C{ars_row}:E{ars_row})")
gt.font = F_VAL_BIG; gt.number_format = "#,##0,,\"M AED\""; gt.border = BORDER; gt.fill = FILL_GREEN_LT

# Same in USD
arsusd_row = ars_row + 1
ws.cell(row=arsusd_row, column=2, value="Our rev-share revenue (USD)").font = F_LBL
ws.cell(row=arsusd_row, column=2).border = BORDER; ws.cell(row=arsusd_row, column=2).fill = FILL_GREEN_LT
for j in range(3):
    col = 3 + j
    cell = ws.cell(row=arsusd_row, column=col, value=f"=C{ars_row}/'1. Cover & Assumptions'!$C$9" if j == 0 else f"={get_column_letter(col)}{ars_row}/'1. Cover & Assumptions'!$C$9")
    cell.font = F_VAL; cell.number_format = "$#,##0"; cell.border = BORDER; cell.fill = FILL_GREEN_LT
gt = ws.cell(row=arsusd_row, column=6, value=f"=SUM(C{arsusd_row}:E{arsusd_row})")
gt.font = F_VAL; gt.number_format = "$#,##0"; gt.border = BORDER; gt.fill = FILL_GREEN_LT

# ── D. TOTAL AIDEOLOGY 3-YEAR REVENUE ──
section_banner(ws, 24, 2, 7, "D.  TOTAL AIDEOLOGY REVENUE  (USD, 3 years)")
th_row2 = 25
for i, h in enumerate(["", "Year 1", "Year 2", "Year 3", "3-Year Total"]):
    c = ws.cell(row=th_row2, column=2 + i, value=h)
    c.font = F_LBL; c.fill = FILL_ACCENT; c.alignment = ALIGN_C; c.border = BORDER

items = [
    ("Fixed fees (Wave 1-4 milestones)", "='8. Revenue Share Model'!C24", "='8. Revenue Share Model'!D24", "='8. Revenue Share Model'!E24"),
    ("Rev-share (USD)",                  f"=C{arsusd_row}",                f"=D{arsusd_row}",                f"=E{arsusd_row}"),
]
for i, (label, *vals) in enumerate(items):
    row = th_row2 + 1 + i
    ws.cell(row=row, column=2, value=label).font = F_LBL; ws.cell(row=row, column=2).border = BORDER
    for j, v in enumerate(vals):
        cell = ws.cell(row=row, column=3 + j, value=v)
        cell.font = F_VAL; cell.number_format = "$#,##0"; cell.border = BORDER
    tot = ws.cell(row=row, column=6, value=f"=SUM(C{row}:E{row})")
    tot.font = F_VAL; tot.number_format = "$#,##0"; tot.border = BORDER

gtr = th_row2 + 1 + len(items)
ws.cell(row=gtr, column=2, value="TOTAL").font = Font(name="Calibri", size=12, bold=True, color=WHITE)
ws.cell(row=gtr, column=2).fill = PatternFill("solid", fgColor=NAVY); ws.cell(row=gtr, column=2).border = BORDER
for j in range(4):
    col = 3 + j; col_letter = get_column_letter(col)
    cell = ws.cell(row=gtr, column=col, value=f"=SUM({col_letter}{th_row2 + 1}:{col_letter}{gtr - 1})")
    cell.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY); cell.number_format = "$#,##0"; cell.border = BORDER
    if col == 6:
        cell.font = Font(name="Calibri", size=14, bold=True, color=WHITE)

# ── E. ACQUISITION & VALUATION ──
section_banner(ws, 31, 2, 7, "E.  ACQUISITION PATHWAY  —  the bigger prize", fill=PatternFill("solid", fgColor="FF6A1B9A"))
ws.merge_cells("B32:G32")
ws["B32"] = "By Year 3-4, e& is paying us 20-28% of revenue forever. Buying us out becomes cheaper than continuing the rev-share."
ws["B32"].font = Font(name="Calibri", size=11, bold=True, color="FF6A1B9A")
ws["B32"].fill = FILL_GOLD; ws["B32"].alignment = ALIGN_L; ws["B32"].border = BORDER
for c in range(3, 8):
    ws.cell(row=32, column=c).fill = FILL_GOLD; ws.cell(row=32, column=c).border = BORDER

# 3-column box: conservative / base / optimistic
summary_box(ws, 33, 2, "Conservative (8x EBITDA)",   "='11. Acquisition & Valuation'!C16", fmt="#,##0,,\"M AED\"", accent=FILL_GOLD)
summary_box(ws, 33, 4, "Base case (multi-telco)",   "='11. Acquisition & Valuation'!D16", fmt="#,##0,,\"M AED\"", accent=FILL_GOLD)
summary_box(ws, 33, 6, "Optimistic (regional moat)","='11. Acquisition & Valuation'!E16", fmt="#,##0,,\"M AED\"", accent=FILL_GOLD)

ws.merge_cells("B36:G36")
ws["B36"] = "These are platform-IP valuations (Year 4 run rate × SaaS multiple). Cash + equity returns beat any consulting margin."
ws["B36"].font = F_DESC; ws["B36"].alignment = ALIGN_L

ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
# =============================================================================
# SHEET 1 — Cover & Assumptions
# =============================================================================
ws = wb.create_sheet("1. Cover & Assumptions")
set_widths(ws, {"A": 3, "B": 44, "C": 16, "D": 10, "E": 60})

ws["B2"] = "AIdeology — Internal Pricing & Staffing Analysis"
ws["B2"].font = F_TITLE
ws["B3"] = "e& B2B SMB Pillar — v0.2 FINAL  ·  aligned to Option A (65/35 declining rev-share, build-then-transfer)"
ws["B3"].font = F_SUB
ws["B4"] = "INTERNAL — NOT FOR e&. Pricing model for AIdeology team only."
ws["B4"].font = Font(name="Calibri", size=10, color=RED, italic=True)
ws["B5"] = "Date: 2026-05-10"
ws["B5"].font = F_NOTE

ws["B7"] = "Global assumptions"
ws["B7"].font = F_H1

assum_hdr = ["Assumption", "Value", "Unit", "Notes / source"]
for i, h in enumerate(assum_hdr):
    c = ws.cell(row=8, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# rows: (label, value, unit, note, fmt)
assum = [
    # ── Programme structure ──
    ("FX (USD → AED)",                       3.67,    "rate",        "Pegged. Used to translate AED rev-share into USD.",                              "0.00"),
    ("Wave 1 duration",                      12,      "weeks",       "M1 SDD → M2 Platform → M3 Customer Agent live → M4 Training/handoff.",          "0"),
    ("Wave 2-3-4 duration (each)",           6,       "weeks",       "Each subsequent wave adds two agents (Wave 4 adds People + platform hardening).","0"),
    ("Working days per week",                5,       "days",        "Standard.",                                                                      "0"),
    ("Hours per working day",                8,       "hours",       "Standard.",                                                                      "0"),
    ("Currency for cost basis",              "USD",   "—",           "AIdeology is USD-denominated; AED used only for e& gross revenue / rev-share.",   "@"),
    # ── Margin & risk ──
    ("AIdeology margin — Wave 1",            0.40,    "%",           "Wave 1 is the heaviest build; agreed fee gives ~40% margin.",                    "0.0%"),
    ("AIdeology margin — Waves 2-4",         0.45,    "%",           "Reusing platform → higher margin justified.",                                   "0.0%"),
    ("Contingency on direct cost",           0.10,    "%",           "Buffer for unknowns (sandbox slip, BSP approval delays, beta iteration).",       "0.0%"),
    ("Risk reserve (held back, internal)",   0.05,    "%",           "Held internally; not in price to e&.",                                          "0.0%"),
    # ── Adoption (per agreement) ──
    ("Customers EOY — Year 1",               24000,   "tenants",     "Per FINAL agreement Y1 forecast.",                                              "#,##0"),
    ("Customers EOY — Year 2",               62000,   "tenants",     "Per FINAL agreement Y2 forecast.",                                              "#,##0"),
    ("Customers EOY — Year 3",               85000,   "tenants",     "Per FINAL agreement Y3 forecast.",                                              "#,##0"),
    ("ARPU AED — Year 1",                    285,     "AED/mo",      "Spark + early Scale mix.",                                                      "#,##0"),
    ("ARPU AED — Year 2",                    350,     "AED/mo",      "Bundle adoption; cross-sell from Customer to Sales/Comms.",                      "#,##0"),
    ("ARPU AED — Year 3",                    420,     "AED/mo",      "Mature mix; Scale + Command tiers dominate.",                                   "#,##0"),
    # ── Rev-share % declining ──
    ("Rev-share % to AIdeology — Year 1",    0.35,    "%",           "Y1-2: 65/35 e&/AIdeology.",                                                      "0%"),
    ("Rev-share % to AIdeology — Year 2",    0.35,    "%",           "Y1-2: 65/35 e&/AIdeology.",                                                      "0%"),
    ("Rev-share % to AIdeology — Year 3",    0.28,    "%",           "Y3: 72/28 (e& AI team takes 40-50% of dev).",                                   "0%"),
    ("Rev-share % to AIdeology — Year 4+",   0.20,    "%",           "Y4+: 80/20 (licensing model).",                                                  "0%"),
    # ── Term ──
    ("Rev-share modelled term",              3,       "years",       "Y1-Y3 modelled; Y4+ shown on Sheet 11 (acquisition).",                          "0"),
]
for i, (label, value, unit, note, fmt) in enumerate(assum):
    r = 9 + i
    ws.cell(row=r, column=2, value=label).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    vc = ws.cell(row=r, column=3, value=value)
    style_input(vc); vc.number_format = fmt
    uc = ws.cell(row=r, column=4, value=unit); uc.font = F_TXT; uc.alignment = ALIGN_C; uc.border = BORDER
    nc = ws.cell(row=r, column=5, value=note); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# Disclaimers
disc_row = 9 + len(assum) + 2
ws.cell(row=disc_row, column=2, value="Important disclaimers").font = F_H1
disclaimers = [
    "• Daily rates in Sheet 2 are AIdeology working assumptions — validate against actual blended rates per role.",
    "• Wave 1 phase staffing (Sheet 3) targets the agreed $1,147,621 fee at ~40% gross margin. Adjust HC/allocation if team capacity changes.",
    "• Waves 2-4 sizing (Sheet 6) assumes the platform from Wave 1 is fully reused — only solution-layer work + integrations remain.",
    "• Subscriber ramp / ARPU on Sheet 8 are the AGREEMENT figures. Sheet 10 stress-tests adoption / ARPU sensitivity.",
    "• e& takes the infrastructure cost (G42 / Azure / OCI) and the LLM / GPU spend per the agreement — those are NOT in our COGS.",
    "• Build-then-transfer: by end of Y2 e& owns agent IP; AIdeology retains the platform. Y3 rev-share drops to 28%; Y4+ to 20% (licensing).",
]
for i, line in enumerate(disclaimers):
    ws.cell(row=disc_row + 1 + i, column=2, value=line).font = F_NOTE


# Programme schedule reference (visual aid)
sch_row = disc_row + 1 + len(disclaimers) + 2
ws.cell(row=sch_row, column=2, value="Programme schedule — at a glance").font = F_H1

# Wave Calendar header
hdr_r = sch_row + 1
hdrs = ["Wave / Milestone", "Weeks", "Fee (USD)", "Cumulative (USD)", "Trigger"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

calendar = [
    ("M1 — SDD, architecture, env",         "W1",          287000,  "Mobilisation."),
    ("M2 — Platform MVP, API, integration", "W2-4",        287000,  "Platform foundations."),
    ("M3 — Customer Agent live (voice/WA)", "W5-8",        287000,  "First agent in production."),
    ("M4 — Training, handoff, docs",        "W9-12",       287000,  "Wave 1 acceptance."),
    ("M5 — Sales Agent design / connectors","W13",         287500,  "Wave 2 kickoff."),
    ("M6 — Sales Agent + Comms Hub live",   "W14-18",      287500,  "Wave 2 acceptance."),
    ("M7 — Finance + Ops design",           "W19",         287500,  "Wave 3 kickoff."),
    ("M8 — Finance + Ops live",             "W20-24",      287500,  "Wave 3 acceptance."),
    ("M9 — People Agent (WPS, payroll)",    "W25",         573000,  "Wave 4 kickoff."),
    ("M10 — Hardening, audit, handoff",     "W26-30",      573000,  "Programme acceptance."),
]
running = 0
for i, (label, weeks, fee, trigger) in enumerate(calendar):
    r = hdr_r + 1 + i
    running += fee
    ws.cell(row=r, column=2, value=label).font = F_LBL; ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=3, value=weeks).font = F_TXT; ws.cell(row=r, column=3).alignment = ALIGN_C; ws.cell(row=r, column=3).border = BORDER
    fc = ws.cell(row=r, column=4, value=fee); style_total(fc); fc.number_format = "$#,##0"
    cc = ws.cell(row=r, column=5, value=running); style_total(cc); cc.number_format = "$#,##0"
    nc = ws.cell(row=r, column=6, value=trigger); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# Total row
tot_r = hdr_r + 1 + len(calendar)
ws.cell(row=tot_r, column=2, value="TOTAL fixed-fee envelope").font = Font(name="Calibri", size=11, bold=True, color=NAVY)
ws.cell(row=tot_r, column=2).fill = FILL_TOTAL; ws.cell(row=tot_r, column=2).border = BORDER
tc = ws.cell(row=tot_r, column=4, value=f"=SUM(D{hdr_r + 1}:D{tot_r - 1})")
style_total(tc); tc.number_format = "$#,##0"; tc.font = Font(name="Calibri", size=11, bold=True, color=NAVY)


# =============================================================================
# SHEET 2 — Roles & Rates
# =============================================================================
ws = wb.create_sheet("2. Roles & Rates")
set_widths(ws, {"A": 3, "B": 44, "C": 14, "D": 18, "E": 18, "F": 70})

ws["B2"] = "Roles & Daily Rates"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: 10 people with different skills. Internal cost = what we pay them per day. Bill rate = what we'd charge for them."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

hdr_r = 5
hdrs = ["Role", "Seniority", "Internal cost / day (USD)", "Bill rate / day (USD)", "Notes"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

roles = [
    ("Engagement / Programme Manager",                     "Senior",    1000, 1750, "Single point of contact to e&; weekly governance; commercial owner."),
    ("Solution Architect (Lead)",                          "Principal", 1100, 2100, "Reference architecture, multi-tenant, latency budget, capability matrix."),
    ("AI / LLM Engineer — Orchestration & Reasoning",      "Senior",     900, 1700, "Agent graph, FSM, tool-calling, memory, guardrails, Portkey routing."),
    ("Voice / Conversational Engineer",                    "Senior",     850, 1650, "Telephony (TF 800), STT/TTS, WhatsApp/Web channels, AR/EN NLP."),
    ("Data & Integration Engineer",                        "Senior",     900, 1550, "Connectors (CRM, calendar, payments), e& billing API, vector store."),
    ("Frontend / UX Engineer",                             "Mid",        750, 1300, "Next.js tenant back office, onboarding, admin console, agent config UI."),
    ("DevOps / Platform Engineer",                         "Senior",     950, 1550, "Helm/Terraform, CI/CD, multi-tenant K8s, observability, cost guardrails."),
    ("QA Engineer",                                        "Mid",        700, 1200, "Test packs, synthetic SMB data, automated regression, voice/chat tests."),
    ("Security / Compliance Lead",                         "Senior",    1050, 1750, "Tenant isolation, NESA / TDRA / ADDA controls, pen-test coordination."),
    ("Customer Success / Beta Ops Lead",                   "Mid",        600,  950, "Beta SMB recruitment & onboarding, feedback loops, runbooks, support."),
]
for i, (role, sen, ic, br, note) in enumerate(roles):
    r = hdr_r + 1 + i
    ws.cell(row=r, column=2, value=role); ws.cell(row=r, column=2).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=3, value=sen); ws.cell(row=r, column=3).font = F_TXT
    ws.cell(row=r, column=3).alignment = ALIGN_C; ws.cell(row=r, column=3).border = BORDER
    icc = ws.cell(row=r, column=4, value=ic); style_input(icc); icc.number_format = "$#,##0"
    brc = ws.cell(row=r, column=5, value=br); style_input(brc); brc.number_format = "$#,##0"
    nc = ws.cell(row=r, column=6, value=note); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

n_roles = len(roles)
role_names = [r[0] for r in roles]

ws.cell(row=hdr_r + 1 + n_roles + 1, column=2, value=f"Total roles in plan: {n_roles}").font = F_LBL
# =============================================================================
# SHEET 3 — Phase Staffing (Wave 1, 12 weeks, 4 phases)
# =============================================================================
ws = wb.create_sheet("3. Phase Staffing")
set_widths(ws, {"A": 3, "B": 50, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14})

ws["B2"] = "Wave 1 — Phase Staffing (12 weeks · 4 phases · 4 milestone payments)"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: How many people work on each phase and how much of their time is on this project. Yellow cells = editable."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

phases = [
    "P1 — Mobilisation & SDD (M1)",
    "P2 — Platform Build & Integration (M2)",
    "P3 — Customer Agent + e& Env Deploy (M3)",
    "P4 — Beta + Training + Handoff (M4)",
]
phase_weeks = [1, 3, 4, 4]   # 12 weeks total
n_phases = len(phases)

# Top calculated table (row 5 headers)
hdr_r = 5
ws.cell(row=hdr_r, column=2, value="Role").font = F_HDR
ws.cell(row=hdr_r, column=2).fill = FILL_HDR; ws.cell(row=hdr_r, column=2).alignment = ALIGN_C; ws.cell(row=hdr_r, column=2).border = BORDER
for i, p in enumerate(phases):
    c = ws.cell(row=hdr_r, column=3 + i, value=p)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER
totp_col = 3 + n_phases  # column for "Total person-weeks"
ws.cell(row=hdr_r, column=totp_col, value="Total p-weeks").font = F_HDR
ws.cell(row=hdr_r, column=totp_col).fill = FILL_HDR; ws.cell(row=hdr_r, column=totp_col).alignment = ALIGN_C
ws.cell(row=hdr_r, column=totp_col).border = BORDER

# row 6 = phase duration
ws.cell(row=6, column=2, value="Phase duration (weeks)").font = F_LBL
ws.cell(row=6, column=2).alignment = ALIGN_L; ws.cell(row=6, column=2).border = BORDER
for i, w in enumerate(phase_weeks):
    cell = ws.cell(row=6, column=3 + i, value=w); style_input(cell); cell.number_format = "0"
ws.cell(row=6, column=totp_col, value="—").font = F_TXT
ws.cell(row=6, column=totp_col).alignment = ALIGN_C; ws.cell(row=6, column=totp_col).border = BORDER

# Editable headcount block starts after the calc table
# Calc table: rows 7..7+n_roles-1 → 7..16
# Total row at 17
HC_HDR_ROW = 21
HC_START_ROW = 22
ALLOC_HDR_ROW = HC_START_ROW + n_roles + 3   # 22 + 10 + 3 = 35
ALLOC_START_ROW = ALLOC_HDR_ROW + 1            # 36

# Sizing: per role per phase  (headcount, allocation%)
# Phases now: P1 (1w), P2 (3w), P3 (4w), P4 (4w)
# Sized for ~8 FTE blended programme team, targeting Wave 1 cost ~$650K (40-45% margin on the agreed $1,147,621).
sizing = {
    # PM (steady throughout)
    0: ((1,1,1,1), (0.8, 0.8, 0.8, 0.8)),
    # Sol Arch (heavy P1-P2, supporting P3, advisory P4)
    1: ((1,1,1,1), (1.0, 1.0, 0.8, 0.5)),
    # AI/LLM Eng (heaviest role; 2 in P2-P3)
    2: ((1,2,2,2), (1.0, 1.0, 1.0, 0.7)),
    # Voice (ramps to P3, supports P4)
    3: ((1,1,1,1), (0.5, 0.9, 1.0, 0.7)),
    # Data/Integ (heavy P2-P3, support P4)
    4: ((1,1,1,1), (0.7, 1.0, 1.0, 0.7)),
    # Frontend (peaks P3-P4)
    5: ((1,1,1,1), (0.5, 0.9, 1.0, 0.8)),
    # DevOps (heavy P2 build-out, steady through P4)
    6: ((1,1,1,1), (0.9, 1.0, 0.9, 0.8)),
    # QA (ramps in P3-P4)
    7: ((0,1,1,1), (0.0, 0.6, 1.0, 1.0)),
    # Security (steady, peaking P4 for handoff)
    8: ((1,1,1,1), (0.6, 0.6, 0.7, 0.7)),
    # CS / Beta Ops (P3-P4)
    9: ((0,0,1,1), (0.0, 0.0, 0.7, 1.0)),
}

# Calculated table — person-weeks per role per phase
for i, role in enumerate(role_names):
    r = 7 + i
    cl = ws.cell(row=r, column=2, value=role); cl.font = F_LBL; cl.alignment = ALIGN_L; cl.border = BORDER
    for j in range(n_phases):
        col = 3 + j
        col_letter = get_column_letter(col)
        hc_row = HC_START_ROW + i
        alloc_row = ALLOC_START_ROW + i
        formula = f"={col_letter}{hc_row}*{col_letter}{alloc_row}*{col_letter}6"
        cell = ws.cell(row=r, column=col, value=formula)
        style_formula(cell); cell.number_format = "0.00"
    tot_cell = ws.cell(row=r, column=totp_col, value=f"=SUM(C{r}:{get_column_letter(totp_col - 1)}{r})")
    style_total(tot_cell); tot_cell.number_format = "0.00"

# Total person-weeks per phase (row 17 = 7 + n_roles)
total_row = 7 + n_roles
ws.cell(row=total_row, column=2, value="Total person-weeks per phase").font = F_TOT
ws.cell(row=total_row, column=2).fill = FILL_TOTAL; ws.cell(row=total_row, column=2).border = BORDER
for j in range(n_phases):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=total_row, column=3 + j, value=f"=SUM({col_letter}7:{col_letter}{total_row - 1})")
    style_total(cell); cell.number_format = "0.00"
gt = ws.cell(row=total_row, column=totp_col, value=f"=SUM(C{total_row}:{get_column_letter(totp_col - 1)}{total_row})")
style_total(gt); gt.number_format = "0.00"; gt.font = Font(name="Calibri", size=10, bold=True, color=NAVY)

# Editable Headcount block
ws.cell(row=HC_HDR_ROW - 2, column=2, value="EDITABLE — Headcount and Allocation % per Role per Phase").font = F_H1
ws.cell(row=HC_HDR_ROW - 1, column=2, value="Headcount per phase (number of people).").font = F_NOTE
ws.cell(row=HC_HDR_ROW, column=2, value="Role").font = F_HDR
ws.cell(row=HC_HDR_ROW, column=2).fill = FILL_HDR; ws.cell(row=HC_HDR_ROW, column=2).alignment = ALIGN_C; ws.cell(row=HC_HDR_ROW, column=2).border = BORDER
for j, p in enumerate(phases):
    c = ws.cell(row=HC_HDR_ROW, column=3 + j, value=p)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

for i, role in enumerate(role_names):
    r = HC_START_ROW + i
    cl = ws.cell(row=r, column=2, value=role); cl.font = F_LBL; cl.alignment = ALIGN_L; cl.border = BORDER
    hc_tuple = sizing[i][0]
    for j in range(n_phases):
        cell = ws.cell(row=r, column=3 + j, value=hc_tuple[j])
        style_input(cell); cell.number_format = "0"

# Editable Allocation block
ws.cell(row=ALLOC_HDR_ROW - 1, column=2, value="Allocation % per phase (share of time on this programme).").font = F_NOTE
ws.cell(row=ALLOC_HDR_ROW, column=2, value="Role").font = F_HDR
ws.cell(row=ALLOC_HDR_ROW, column=2).fill = FILL_HDR; ws.cell(row=ALLOC_HDR_ROW, column=2).alignment = ALIGN_C; ws.cell(row=ALLOC_HDR_ROW, column=2).border = BORDER
for j, p in enumerate(phases):
    c = ws.cell(row=ALLOC_HDR_ROW, column=3 + j, value=p)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

for i, role in enumerate(role_names):
    r = ALLOC_START_ROW + i
    cl = ws.cell(row=r, column=2, value=role); cl.font = F_LBL; cl.alignment = ALIGN_L; cl.border = BORDER
    al_tuple = sizing[i][1]
    for j in range(n_phases):
        cell = ws.cell(row=r, column=3 + j, value=al_tuple[j])
        style_input(cell); cell.number_format = "0%"


# =============================================================================
# SHEET 4 — Phase Effort & Cost (Wave 1)
# =============================================================================
ws = wb.create_sheet("4. Phase Effort & Cost")
set_widths(ws, {"A": 3, "B": 44, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 16, "I": 12})

ws["B2"] = "Wave 1 — Internal Cost per Phase"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: 'how long they work' (Sheet 3) × 'their daily rate' (Sheet 2). Total at the bottom = our Wave 1 personnel cost."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

ws["B5"] = "INTERNAL COST (fully-loaded AIdeology cost)"
ws["B5"].font = F_H1

hdr_r = 6
hdr = ["Role", "Cost / day"] + phases + ["Total cost", "% of total"]
for i, h in enumerate(hdr):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

WORKING_DAYS_REF = "'1. Cover & Assumptions'!$C$12"   # Working days per week — row 12 of assum block (FX=9, Wave1=10, Wave2-4=11, WorkDays=12)
# Verify: assum starts row 9; FX is row 9; Wave 1 dur is row 10; Wave 2-4 dur is row 11; Working days = row 12.

for i, role in enumerate(role_names):
    r = hdr_r + 1 + i
    cl = ws.cell(row=r, column=2, value=role); cl.font = F_LBL; cl.alignment = ALIGN_L; cl.border = BORDER
    sheet2_row = 6 + i
    cell = ws.cell(row=r, column=3, value=f"='2. Roles & Rates'!D{sheet2_row}")
    style_formula(cell); cell.number_format = "$#,##0"
    sheet3_row = 7 + i
    for j in range(n_phases):
        sh3_col = get_column_letter(3 + j)            # C..F
        col_letter = get_column_letter(4 + j)          # D..G (output cols)
        formula = f"='3. Phase Staffing'!{sh3_col}{sheet3_row}*{WORKING_DAYS_REF}*$C{r}"
        c = ws.cell(row=r, column=4 + j, value=formula)
        style_formula(c); c.number_format = "$#,##0"
    tot_col = 4 + n_phases    # column H
    tot = ws.cell(row=r, column=tot_col, value=f"=SUM(D{r}:{get_column_letter(tot_col - 1)}{r})")
    style_total(tot); tot.number_format = "$#,##0"
    pct_col = tot_col + 1
    pct = ws.cell(row=r, column=pct_col, value=f"=IFERROR({get_column_letter(tot_col)}{r}/${get_column_letter(tot_col)}${hdr_r + n_roles + 1},0)")
    style_formula(pct); pct.number_format = "0.0%"

# Total row
tr = hdr_r + 1 + n_roles
ws.cell(row=tr, column=2, value="Total internal cost per phase").font = F_TOT
ws.cell(row=tr, column=2).fill = FILL_TOTAL; ws.cell(row=tr, column=2).border = BORDER
ws.cell(row=tr, column=3, value="").border = BORDER; ws.cell(row=tr, column=3).fill = FILL_TOTAL
for j in range(n_phases):
    col_letter = get_column_letter(4 + j)
    cell = ws.cell(row=tr, column=4 + j, value=f"=SUM({col_letter}{hdr_r + 1}:{col_letter}{tr - 1})")
    style_total(cell); cell.number_format = "$#,##0"
tot_col = 4 + n_phases   # H
gt = ws.cell(row=tr, column=tot_col, value=f"=SUM({get_column_letter(tot_col)}{hdr_r + 1}:{get_column_letter(tot_col)}{tr - 1})")
style_total(gt); gt.number_format = "$#,##0"; gt.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
ws.cell(row=tr, column=tot_col + 1, value="100%").font = F_TOT
ws.cell(row=tr, column=tot_col + 1).fill = FILL_TOTAL; ws.cell(row=tr, column=tot_col + 1).border = BORDER
ws.cell(row=tr, column=tot_col + 1).alignment = ALIGN_R; ws.cell(row=tr, column=tot_col + 1).number_format = "0%"

# Save the cost reference for Sheet 7
W1_COST_CELL = f"'4. Phase Effort & Cost'!$H${tr}"   # H{17} = personnel cost total
W1_COST_ROW = tr
# =============================================================================
# SHEET 5 — Other Costs (Wave 1)
# =============================================================================
ws = wb.create_sheet("5. Other Costs")
set_widths(ws, {"A": 3, "B": 44, "C": 14, "D": 14, "E": 16, "F": 14, "G": 60})

ws["B2"] = "Wave 1 — Non-People Costs"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: Everything we spend that isn't salaries — cloud, AI APIs, WhatsApp BSP, security, travel."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

hdr = ["Cost item", "Unit cost", "Quantity / months", "Total cost", "Phase trigger", "Notes"]
for i, h in enumerate(hdr):
    c = ws.cell(row=5, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# 12 weeks ≈ 3 months  →  qty = 3 for monthly run-rate items
others = [
    ("Cloud / K8s dev + staging cluster (e& OCI / Azure CSP)",  6000, 3, "P1-P4", "Multi-tenant cluster footprint during build & beta. Production sized separately."),
    ("LLM Gateway licence (Portkey / equivalent)",              1500, 3, "P1-P4", "Routing, fallback, observability, cost guardrails across Azure OpenAI / Anthropic."),
    ("LLM API spend — build, alpha, beta",                      4000, 3, "P2-P4", "Prompt engineering, regression test runs, beta tenant traffic."),
    ("STT / TTS API (Deepgram, ElevenLabs / Azure Speech)",     2500, 3, "P2-P4", "Voice agent training & beta. Arabic/English, low-latency."),
    ("WhatsApp BSP (e& or partner) — sandbox + tests",          1500, 3, "P2-P4", "Template approval, opt-in flows, message-rate tests on UAE numbers."),
    ("Observability stack (Grafana Cloud, Loki, Langfuse)",     1500, 3, "P1-P4", "Hosted observability for AIdeology team; production runs in e& env."),
    ("CI/CD & repo hosting (GitHub Enterprise)",                 800, 3, "P1-P4", "Pipelines, secrets, environments, security scanning."),
    ("Security tooling (SAST/DAST/SCA — Snyk / Checkmarx)",     2000, 3, "P1-P4", "Continuous static & dependency scanning."),
    ("Pen-test (third-party, before handoff)",                  25000, 1, "P4",   "External pen-test prior to production handoff."),
    ("Travel — UAE (3 people × 5 days × 3 trips)",              9000, 3, "P1-P4", "On-site mobilisation, integration kickoff, handoff."),
    ("Marketing / launch event support",                         8000, 1, "P4",   "Handoff event + collateral + demo videos."),
    ("Misc (printing, demos, contingency petty)",               1500, 3, "All",   "Petty cash for the engagement."),
]
for i, (item, unit, qty, trigger, note) in enumerate(others):
    r = 6 + i
    ws.cell(row=r, column=2, value=item).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    uc = ws.cell(row=r, column=3, value=unit); style_input(uc); uc.number_format = "$#,##0"
    qc = ws.cell(row=r, column=4, value=qty); style_input(qc); qc.number_format = "0"
    tc = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); style_total(tc); tc.number_format = "$#,##0"
    pc = ws.cell(row=r, column=6, value=trigger); pc.font = F_TXT; pc.alignment = ALIGN_C; pc.border = BORDER
    nc = ws.cell(row=r, column=7, value=note); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

W1_OTHER_TOT_ROW = 6 + len(others)
ws.cell(row=W1_OTHER_TOT_ROW, column=2, value="Total other costs (subtotal)").font = F_TOT
ws.cell(row=W1_OTHER_TOT_ROW, column=2).fill = FILL_TOTAL; ws.cell(row=W1_OTHER_TOT_ROW, column=2).border = BORDER
ws.cell(row=W1_OTHER_TOT_ROW, column=5, value=f"=SUM(E6:E{W1_OTHER_TOT_ROW - 1})")
style_total(ws.cell(row=W1_OTHER_TOT_ROW, column=5))
ws.cell(row=W1_OTHER_TOT_ROW, column=5).number_format = "$#,##0"
ws.cell(row=W1_OTHER_TOT_ROW, column=5).font = Font(name="Calibri", size=11, bold=True, color=NAVY)

W1_OTHER_REF = f"'5. Other Costs'!$E${W1_OTHER_TOT_ROW}"


# =============================================================================
# SHEET 6 — Wave 2-4 Build Costs (3 waves, post-platform)
# =============================================================================
ws = wb.create_sheet("6. Wave 2-4 Build Costs")
set_widths(ws, {"A": 3, "B": 38, "C": 16, "D": 22, "E": 22, "F": 22, "G": 18, "H": 50})

ws["B2"] = "Waves 2-4 — Cost of the Next 5 Agents (post-platform)"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: After Wave 1 the platform is built. Wave 2 adds Sales+Comms (6w). Wave 3 adds Finance+Ops (6w). Wave 4 adds People + production hardening (6w)."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

# Wave structure for the agreement: 3 waves, each 6 weeks
waves = [
    ("Wave 2 — Sales + Comms (W13-18)",            6),
    ("Wave 3 — Finance + Ops (W19-24)",            6),
    ("Wave 4 — People + Hardening (W25-30)",       6),
]
n_waves = len(waves)

hdr_r = 5
hdrs = ["Role", "Cost / day"] + [w[0] for w in waves] + ["Total p-weeks", "Notes"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Person-weeks per role per wave.
# Wave 4 is intentionally heavier than Wave 2/3 — it carries the People Agent
# AND the platform hardening / security audit / pen-test / final handoff scope.
wave_pweeks = {
    0: [3, 3, 4],            # PM
    1: [4, 4, 5],            # Sol Arch (lead/QA architecture per wave)
    2: [12, 10, 10],         # AI/LLM Eng (heaviest each wave; W4 hardening adds eval work)
    3: [8, 0, 0],            # Voice (heavy in Wave 2 Comms; none in Wave 3-4)
    4: [10, 14, 14],         # Data/Integ (Wave 3 = Finance/Ops APIs; Wave 4 = HR systems + audit hooks)
    5: [6, 6, 7],            # Frontend
    6: [3, 3, 10],           # DevOps (Wave 4 hardening peak — observability, scaling, runbooks)
    7: [4, 4, 10],           # QA (Wave 4 full regression + acceptance suite)
    8: [3, 4, 10],           # Security (Wave 4 = audit, pen-test prep, NESA controls)
    9: [3, 3, 5],            # CS / Beta Ops
}

note_per_role = {
    0: "Engagement coordination; lighter than Wave 1.",
    1: "Architect supports SDD per wave; capacity check with platform.",
    2: "Heaviest role per agent — orchestration, prompts, tools, evals.",
    3: "Comms Hub = heaviest voice; Finance / Ops / People → minimal voice.",
    4: "Connectors are the wave-driver: Finance APIs (FTA/banks); HR systems (MoHRE/WPS).",
    5: "Tenant UI tweaks per agent + admin console.",
    6: "Per-wave deployment, observability hooks, cost guardrails — Wave 4 hardening peak.",
    7: "Test packs, regression, acceptance per agent.",
    8: "Higher for Finance & People (PII, payroll regs); audit at Wave 4.",
    9: "Pilot tenant onboarding & support per wave.",
}

for i, role in enumerate(role_names):
    r = hdr_r + 1 + i
    cl = ws.cell(row=r, column=2, value=role); cl.font = F_LBL; cl.alignment = ALIGN_L; cl.border = BORDER
    sheet2_row = 6 + i
    cd = ws.cell(row=r, column=3, value=f"='2. Roles & Rates'!D{sheet2_row}")
    style_formula(cd); cd.number_format = "$#,##0"
    pweeks_list = wave_pweeks[i]
    for j, pw in enumerate(pweeks_list):
        cell = ws.cell(row=r, column=4 + j, value=pw)
        style_input(cell); cell.number_format = "0"
    tot_col = 4 + n_waves    # column G (=7)
    tot = ws.cell(row=r, column=tot_col, value=f"=SUM(D{r}:{get_column_letter(tot_col - 1)}{r})")
    style_total(tot); tot.number_format = "0"
    nc = ws.cell(row=r, column=tot_col + 1, value=note_per_role[i]); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# Person-weeks subtotal row
tr_pw = hdr_r + 1 + n_roles
ws.cell(row=tr_pw, column=2, value="Total person-weeks per wave").font = F_TOT
ws.cell(row=tr_pw, column=2).fill = FILL_TOTAL; ws.cell(row=tr_pw, column=2).border = BORDER
for j in range(n_waves):
    col_letter = get_column_letter(4 + j)
    cell = ws.cell(row=tr_pw, column=4 + j, value=f"=SUM({col_letter}{hdr_r + 1}:{col_letter}{tr_pw - 1})")
    style_total(cell); cell.number_format = "0"
tot_col = 4 + n_waves
gtot = ws.cell(row=tr_pw, column=tot_col, value=f"=SUM({get_column_letter(tot_col)}{hdr_r + 1}:{get_column_letter(tot_col)}{tr_pw - 1})")
style_total(gtot); gtot.number_format = "0"

# Personnel cost per wave (USD) section
cr_label_row = tr_pw + 2
ws.cell(row=cr_label_row, column=2, value="Personnel cost per wave (USD)").font = F_H1
hdr_r2 = cr_label_row + 1
hdrs2 = ["Item"] + [w[0] for w in waves] + ["Total"]
for i, h in enumerate(hdrs2):
    c = ws.cell(row=hdr_r2, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

WD_REF = WORKING_DAYS_REF
pers_row = hdr_r2 + 1
ws.cell(row=pers_row, column=2, value="Personnel cost").font = F_LBL
ws.cell(row=pers_row, column=2).alignment = ALIGN_L; ws.cell(row=pers_row, column=2).border = BORDER
for j in range(n_waves):
    col_pw = get_column_letter(4 + j)
    parts = []
    for i in range(n_roles):
        r_role = hdr_r + 1 + i
        parts.append(f"{col_pw}{r_role}*$C{r_role}")
    formula = f"=({'+'.join(parts)})*{WD_REF}"
    cell = ws.cell(row=pers_row, column=4 + j, value=formula)
    style_total(cell); cell.number_format = "$#,##0"
gt = ws.cell(row=pers_row, column=4 + n_waves, value=f"=SUM(D{pers_row}:{get_column_letter(3 + n_waves)}{pers_row})")
style_total(gt); gt.number_format = "$#,##0"; gt.font = Font(name="Calibri", size=11, bold=True, color=NAVY)

# Other costs per wave (~25% of Wave 1 envelope per wave is conservative)
oc_row = pers_row + 1
ws.cell(row=oc_row, column=2, value="Other costs (per wave, ~25% of Wave 1 envelope)").font = F_LBL
ws.cell(row=oc_row, column=2).alignment = ALIGN_L; ws.cell(row=oc_row, column=2).border = BORDER
for j in range(n_waves):
    cell = ws.cell(row=oc_row, column=4 + j, value=f"={W1_OTHER_REF}*0.25")
    style_formula(cell); cell.number_format = "$#,##0"
gt2 = ws.cell(row=oc_row, column=4 + n_waves, value=f"=SUM(D{oc_row}:{get_column_letter(3 + n_waves)}{oc_row})")
style_total(gt2); gt2.number_format = "$#,##0"

# Total wave cost
tw_row = oc_row + 1
ws.cell(row=tw_row, column=2, value="Total fully-loaded cost per wave").font = F_TOT
ws.cell(row=tw_row, column=2).fill = FILL_TOTAL; ws.cell(row=tw_row, column=2).border = BORDER
for j in range(n_waves):
    col_letter = get_column_letter(4 + j)
    cell = ws.cell(row=tw_row, column=4 + j, value=f"={col_letter}{pers_row}+{col_letter}{oc_row}")
    style_total(cell); cell.number_format = "$#,##0"
gt3 = ws.cell(row=tw_row, column=4 + n_waves, value=f"=SUM(D{tw_row}:{get_column_letter(3 + n_waves)}{tw_row})")
style_total(gt3); gt3.number_format = "$#,##0"; gt3.font = Font(name="Calibri", size=11, bold=True, color=NAVY)

# Save references for Sheet 7
W2_4_TW_ROW = tw_row     # row in Sheet 6 with total wave cost
# =============================================================================
# SHEET 7 — Pricing Summary (cost vs AGREED fees → effective margin)
# =============================================================================
ws = wb.create_sheet("7. Pricing Summary")
set_widths(ws, {"A": 3, "B": 38, "C": 18, "D": 18, "E": 18, "F": 16, "G": 60})

ws["B2"] = "Pricing Summary — Agreed Fees vs Our Internal Cost"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: e& has agreed to pay $3,443,621 across 4 waves. This sheet checks our cost basis lands inside the fee — the gap is our margin."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

# Table: Wave | Agreed fee (USD) | Internal cost | Margin USD | Margin %
ws["B5"] = "Wave-by-wave economics"
ws["B5"].font = F_H1

hdr_r = 6
for i, h in enumerate(["Wave", "Agreed fee (USD)", "Internal cost (USD)", "Margin (USD)", "Margin %", "Notes"]):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Row 7 = header  → data rows 9..12 give per-wave breakdown
# Use rows 9-12 to keep round references
# Wave 1 cost = Sheet 4 personnel + Sheet 5 other
W1_PERS = W1_COST_CELL                     # personnel total
W1_OTHER = W1_OTHER_REF
W1_COST_FORMULA = f"={W1_PERS}+{W1_OTHER}"

# Row layout
data_start = hdr_r + 1   # row 7
wave_rows = [
    ("Wave 1 — Platform + Customer Agent",         1147621, W1_COST_FORMULA,                                       "12 weeks · Milestones M1-M4 · 4×$287k."),
    ("Wave 2 — Sales + Comms",                      575000, f"='6. Wave 2-4 Build Costs'!D{W2_4_TW_ROW}",          "6 weeks · Milestones M5-M6 · 2×$287.5k."),
    ("Wave 3 — Finance + Ops",                      575000, f"='6. Wave 2-4 Build Costs'!E{W2_4_TW_ROW}",          "6 weeks · Milestones M7-M8 · 2×$287.5k."),
    ("Wave 4 — People + Hardening",                1146000, f"='6. Wave 2-4 Build Costs'!F{W2_4_TW_ROW}",          "6 weeks · Milestones M9-M10 · 2×$573k."),
]

for i, (label, fee, cost_formula, note) in enumerate(wave_rows):
    r = data_start + i        # rows 7, 8, 9, 10? Actually want 9,10,11,12
    pass
# Use explicit row numbers 9..12 to keep readable references
DATA_ROW_W1 = 9
DATA_ROW_W2 = 10
DATA_ROW_W3 = 11
DATA_ROW_W4 = 12
DATA_ROW_TOTAL = 13

# Optional sub-header for clarity (row 7 spacer / row 8 mini-note)
sh = ws.cell(row=7, column=2, value="Wave-level table (agreed fee comes from the FINAL agreement; cost is computed from Sheets 4-6).")
sh.font = F_NOTE; ws.merge_cells("B7:G7"); sh.alignment = ALIGN_L

for i, (label, fee, cost_formula, note) in enumerate(wave_rows):
    r = DATA_ROW_W1 + i
    ws.cell(row=r, column=2, value=label).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    fc = ws.cell(row=r, column=3, value=fee); style_total(fc); fc.number_format = "$#,##0"
    cc = ws.cell(row=r, column=4, value=cost_formula); style_formula(cc); cc.number_format = "$#,##0"
    mc = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); style_total(mc); mc.number_format = "$#,##0"
    mp = ws.cell(row=r, column=6, value=f"=IFERROR((C{r}-D{r})/C{r},0)"); style_formula(mp); mp.number_format = "0.0%"
    nc = ws.cell(row=r, column=7, value=note); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# Total row
r = DATA_ROW_TOTAL
ws.cell(row=r, column=2, value="TOTAL — programme envelope").font = Font(name="Calibri", size=11, bold=True, color=NAVY)
ws.cell(row=r, column=2).fill = FILL_TOTAL; ws.cell(row=r, column=2).border = BORDER
for col in [3, 4, 5]:
    col_letter = get_column_letter(col)
    cell = ws.cell(row=r, column=col, value=f"=SUM({col_letter}{DATA_ROW_W1}:{col_letter}{DATA_ROW_W4})")
    style_total(cell); cell.number_format = "$#,##0"
    cell.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
mp = ws.cell(row=r, column=6, value=f"=IFERROR((C{r}-D{r})/C{r},0)")
style_total(mp); mp.number_format = "0.0%"; mp.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
ws.cell(row=r, column=7, value="Total fixed-fee envelope. Margin gap = our cushion.").font = F_NOTE
ws.cell(row=r, column=7).border = BORDER

# Risk reserve (held internally, not in price)
rr_row = DATA_ROW_TOTAL + 2
ws.cell(row=rr_row, column=2, value="Risk reserve (held internally, off-fee)").font = F_LBL
ws.cell(row=rr_row, column=2).alignment = ALIGN_L; ws.cell(row=rr_row, column=2).border = BORDER
rrc = ws.cell(row=rr_row, column=3, value="=D13*'1. Cover & Assumptions'!$C$18")
style_formula(rrc); rrc.number_format = "$#,##0"
ws.cell(row=rr_row, column=7, value="5% of total cost held at AIdeology level for genuine risk events.").font = F_NOTE
ws.cell(row=rr_row, column=7).border = BORDER

# Margin reality check
mr_row = rr_row + 2
ws.cell(row=mr_row, column=2, value="Margin reality check").font = F_H1

bullets = [
    "• Targets in Sheet 1 are 40% (Wave 1) and 45% (Waves 2-4). The agreed fees are generous: actual computed margins are HIGHER (see column F).",
    "• The 'extra' margin on top of target is our cushion against scope creep, sandbox slip, BSP delays, and subscriber-ramp shortfalls.",
    "• Wave 4 is the heaviest engineering wave (hardening, pen-test, audit, full handoff). Cost overruns there are the main risk.",
    "• Even if we under-deliver Wave 1 by 20%, programme gross margin stays > 50%. The structure is robust.",
    "• The bigger-picture economics — rev-share to Year 3 and acquisition valuation — are on Sheets 8 and 11.",
]
for i, b in enumerate(bullets):
    c = ws.cell(row=mr_row + 1 + i, column=2, value=b); c.font = F_NOTE; c.alignment = ALIGN_L
    ws.merge_cells(start_row=mr_row + 1 + i, start_column=2, end_row=mr_row + 1 + i, end_column=7)


# =============================================================================
# SHEET 8 — Revenue Share Model (3-Year View)
# =============================================================================
ws = wb.create_sheet("8. Revenue Share Model")
set_widths(ws, {"A": 3, "B": 44, "C": 18, "D": 18, "E": 18, "F": 18, "G": 50})

ws["B2"] = "Revenue Share — Our Recurring Income (per FINAL agreement)"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: SMBs pay e& monthly. We get a % of e&'s SMB revenue. The % declines as e& takes over delivery: 35% Y1-2, 28% Y3, 20% Y4+."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

# A. Adoption block
ws["B5"] = "A. SMB customer ramp (per agreement)"
ws["B5"].font = F_H1

hdr_r = 6
for i, h in enumerate(["Metric", "Year 1", "Year 2", "Year 3", "", "Notes"]):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# rows 7..10
# Customers EOY
ws.cell(row=7, column=2, value="Customers — end of year").font = F_LBL
ws.cell(row=7, column=2).alignment = ALIGN_L; ws.cell(row=7, column=2).border = BORDER
for j, asum_row in enumerate([19, 20, 21]):   # Cover assumptions rows for Customers EOY Y1/Y2/Y3
    cell = ws.cell(row=7, column=3 + j, value=f"='1. Cover & Assumptions'!$C${asum_row}")
    style_formula(cell); cell.number_format = "#,##0"
ws.cell(row=7, column=7, value="EOY headcount per FINAL agreement.").font = F_NOTE; ws.cell(row=7, column=7).border = BORDER

# Avg paying customers — editable, defaulted to agreement-aligned values
# Defaults: 10,672 / 36,667 / 75,000  →  produce agreement gross of AED 36.5M / 154M / 378M
ws.cell(row=8, column=2, value="Avg paying customers (year)").font = F_LBL
ws.cell(row=8, column=2).alignment = ALIGN_L; ws.cell(row=8, column=2).border = BORDER
for j, default in enumerate([10672, 36667, 75000]):
    cell = ws.cell(row=8, column=3 + j, value=default)
    style_input(cell); cell.number_format = "#,##0"
ws.cell(row=8, column=7, value="Defaults match the FINAL agreement (slow start in Y1; ramp accelerates Y2-Y3). Editable.").font = F_NOTE
ws.cell(row=8, column=7).border = BORDER

# ARPU AED
ws.cell(row=9, column=2, value="ARPU (AED / month)").font = F_LBL
ws.cell(row=9, column=2).alignment = ALIGN_L; ws.cell(row=9, column=2).border = BORDER
for j, asum_row in enumerate([22, 23, 24]):  # Cover assumptions ARPU rows
    cell = ws.cell(row=9, column=3 + j, value=f"='1. Cover & Assumptions'!$C${asum_row}")
    style_formula(cell); cell.number_format = "#,##0\" AED\""
ws.cell(row=9, column=7, value="From Sheet 1.").font = F_NOTE; ws.cell(row=9, column=7).border = BORDER

# Effective annual ARPU AED
ws.cell(row=10, column=2, value="Annual ARPU per customer (AED)").font = F_LBL
ws.cell(row=10, column=2).alignment = ALIGN_L; ws.cell(row=10, column=2).border = BORDER
for j in range(3):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=10, column=3 + j, value=f"={col_letter}9*12")
    style_formula(cell); cell.number_format = "#,##0"
ws.cell(row=10, column=7, value="Monthly ARPU × 12.").font = F_NOTE; ws.cell(row=10, column=7).border = BORDER

# B. Gross e& revenue & rev-share
GROSS_ROW = 11
ws.cell(row=GROSS_ROW, column=2, value="e& gross SaaS revenue (AED / year)").font = F_TOT
ws.cell(row=GROSS_ROW, column=2).fill = FILL_TOTAL; ws.cell(row=GROSS_ROW, column=2).border = BORDER
for j in range(3):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=GROSS_ROW, column=3 + j, value=f"={col_letter}8*{col_letter}10")
    style_total(cell); cell.number_format = "#,##0"
ws.cell(row=GROSS_ROW, column=7, value="Avg customers × annual ARPU.").font = F_NOTE
ws.cell(row=GROSS_ROW, column=7).border = BORDER

# Rev-share % (declining)
RSPCT_ROW = 12
ws.cell(row=RSPCT_ROW, column=2, value="Rev-share % to AIdeology (declining)").font = F_LBL
ws.cell(row=RSPCT_ROW, column=2).alignment = ALIGN_L; ws.cell(row=RSPCT_ROW, column=2).border = BORDER
for j, asum_row in enumerate([25, 26, 27]):  # Y1/Y2/Y3 rev-share %
    cell = ws.cell(row=RSPCT_ROW, column=3 + j, value=f"='1. Cover & Assumptions'!$C${asum_row}")
    style_formula(cell); cell.number_format = "0%"
ws.cell(row=RSPCT_ROW, column=7, value="35% Y1-Y2, 28% Y3 (build-then-transfer).").font = F_NOTE
ws.cell(row=RSPCT_ROW, column=7).border = BORDER

# AIdeology rev-share AED
RSREV_ROW = 13
ws.cell(row=RSREV_ROW, column=2, value="AIdeology rev-share revenue (AED / year)").font = F_TOT
ws.cell(row=RSREV_ROW, column=2).fill = FILL_TOTAL; ws.cell(row=RSREV_ROW, column=2).border = BORDER
for j in range(3):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=RSREV_ROW, column=3 + j, value=f"={col_letter}{GROSS_ROW}*{col_letter}{RSPCT_ROW}")
    style_total(cell); cell.number_format = "#,##0"
ws.cell(row=RSREV_ROW, column=7, value="Gross × rev-share %.").font = F_NOTE
ws.cell(row=RSREV_ROW, column=7).border = BORDER

# AIdeology rev-share USD (using FX from Sheet 1 row 9)
RSUSD_ROW = 14
ws.cell(row=RSUSD_ROW, column=2, value="AIdeology rev-share revenue (USD / year)").font = F_TOT
ws.cell(row=RSUSD_ROW, column=2).fill = FILL_TOTAL; ws.cell(row=RSUSD_ROW, column=2).border = BORDER
for j in range(3):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=RSUSD_ROW, column=3 + j, value=f"={col_letter}{RSREV_ROW}/'1. Cover & Assumptions'!$C$9")
    style_total(cell); cell.number_format = "$#,##0"
ws.cell(row=RSUSD_ROW, column=7, value="AED ÷ FX (3.67).").font = F_NOTE
ws.cell(row=RSUSD_ROW, column=7).border = BORDER

# Fees recognised by year
ws.cell(row=16, column=2, value="B. Fixed fees recognised by year").font = F_H1
hdr_b = 17
for i, h in enumerate(["Item", "Year 1", "Year 2", "Year 3", "", "Notes"]):
    c = ws.cell(row=hdr_b, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Wave 1 (12 wks) + Wave 2 (6 wks) + Wave 3 (6 wks) = 24 weeks ≈ 6 months → all in Y1
# Wave 4 (6 wks) starting W25 = in Y1 if calendar starts at M0 of Y1 = end of Y1
# All four waves = 30 weeks = ~7 months → all Y1
fee_rows = [
    ("Wave 1 — Platform + Customer Agent",  1147621, 0, 0, "Recognised entirely in Y1 (M1-M4)."),
    ("Wave 2 — Sales + Comms",                575000, 0, 0, "Recognised in Y1 (M5-M6)."),
    ("Wave 3 — Finance + Ops",                575000, 0, 0, "Recognised in Y1 (M7-M8)."),
    ("Wave 4 — People + Hardening",          1146000, 0, 0, "Recognised in Y1 (M9-M10)."),
]
for i, (label, y1, y2, y3, note) in enumerate(fee_rows):
    r = hdr_b + 1 + i
    ws.cell(row=r, column=2, value=label).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    for j, v in enumerate([y1, y2, y3]):
        cell = ws.cell(row=r, column=3 + j, value=v)
        style_input(cell); cell.number_format = "$#,##0"
    nc = ws.cell(row=r, column=7, value=note); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

FEE_SUB_ROW = hdr_b + 1 + len(fee_rows)
ws.cell(row=FEE_SUB_ROW, column=2, value="Subtotal — fees recognised").font = F_TOT
ws.cell(row=FEE_SUB_ROW, column=2).fill = FILL_TOTAL; ws.cell(row=FEE_SUB_ROW, column=2).border = BORDER
for j in range(3):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=FEE_SUB_ROW, column=3 + j, value=f"=SUM({col_letter}{hdr_b + 1}:{col_letter}{FEE_SUB_ROW - 1})")
    style_total(cell); cell.number_format = "$#,##0"

# C. Total AIdeology revenue per year (USD)
ws.cell(row=23, column=2, value="C. Total AIdeology revenue per year (USD)").font = F_H1
hdr_c = 24
for i, h in enumerate(["Item", "Year 1", "Year 2", "Year 3", "3-Year Total"]):
    c = ws.cell(row=hdr_c, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Row 24 contains "Item Y1 Y2 Y3 3-Year Total" header
# Data rows 25-28
ws.cell(row=24, column=2, value="Fixed fees")  # actually we need to place data starting row 25
# ↑ overwrite. Let me just use rows 25..27 cleanly.
# Actually hdr_c = 24, so data rows start at 25.
# Re-set headers explicitly above.

# Data: row 25 = Fixed fees, row 26 = Rev-share, row 27 = TOTAL
# Total row goes to ws[Sheet 0] reference. Make sure these row numbers are correct.

# Fixed fees row
r = 25
ws.cell(row=r, column=2, value="Fixed fees").font = F_LBL; ws.cell(row=r, column=2).border = BORDER
for j in range(3):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=r, column=3 + j, value=f"={col_letter}{FEE_SUB_ROW}")
    style_formula(cell); cell.number_format = "$#,##0"
tot = ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})")
style_total(tot); tot.number_format = "$#,##0"

# Rev-share row
r = 26
ws.cell(row=r, column=2, value="Rev-share (USD)").font = F_LBL; ws.cell(row=r, column=2).border = BORDER
for j in range(3):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=r, column=3 + j, value=f"={col_letter}{RSUSD_ROW}")
    style_formula(cell); cell.number_format = "$#,##0"
tot = ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})")
style_total(tot); tot.number_format = "$#,##0"

# Total row
r = 27
ws.cell(row=r, column=2, value="TOTAL AIdeology revenue").font = Font(name="Calibri", size=11, bold=True, color=WHITE)
ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=NAVY); ws.cell(row=r, column=2).border = BORDER
for j in range(3):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=r, column=3 + j, value=f"=SUM({col_letter}25:{col_letter}26)")
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY); cell.number_format = "$#,##0"; cell.border = BORDER
gt = ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})")
gt.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
gt.fill = PatternFill("solid", fgColor=NAVY); gt.number_format = "$#,##0"; gt.border = BORDER

# Note
ws.merge_cells("B29:G29")
ws.cell(row=29, column=2, value="Note: All 10 milestones (Wave 1-4) close inside Year 1 since the build runs ~30 weeks. Y2 and Y3 = pure rev-share.").font = F_NOTE
ws.cell(row=29, column=2).alignment = ALIGN_L
# =============================================================================
# SHEET 9 — Payment Milestones (Wave 1-4, 10 milestones)
# =============================================================================
ws = wb.create_sheet("9. Payment Milestones")
set_widths(ws, {"A": 3, "B": 8, "C": 38, "D": 14, "E": 12, "F": 18, "G": 50})

ws["B2"] = "Payment Milestones — When Does e& Pay Us?"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: 10 milestones over 30 weeks. Each milestone has an acceptance test before payment."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

hdr_r = 5
for i, h in enumerate(["#", "Milestone / trigger", "Wave", "Week", "Amount (USD)", "Acceptance criteria"]):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

milestones = [
    (1,  "SDD, architecture, env setup",                      "Wave 1", "W1",   287000, "SDD signed off · target architecture review accepted · environments stood up."),
    (2,  "Platform MVP, API framework, integration layer",    "Wave 1", "W4",   287000, "Multi-tenant platform spun up · API layer live · core integrations stubbed."),
    (3,  "Customer Agent live (voice + WhatsApp)",            "Wave 1", "W8",   287000, "Customer Agent in production for pilot tenants · voice + WhatsApp + web channels."),
    (4,  "Training, handoff, documentation",                  "Wave 1", "W12",  287000, "e& team can independently deploy/customise · documentation accepted."),
    (5,  "Sales Agent design + Comms Hub design",             "Wave 2", "W13",  287500, "SDDs accepted · CRM/lead-scoring stubs working · channel router scaffolded."),
    (6,  "Sales Agent + Comms Hub live",                      "Wave 2", "W18",  287500, "Both agents in production · CRM/Comms integrations validated · acceptance tests passed."),
    (7,  "Finance + Ops Agents design",                       "Wave 3", "W19",  287500, "FTA / VAT compliance design accepted · Ops SOP KB schema validated."),
    (8,  "Finance + Ops Agents live",                         "Wave 3", "W24",  287500, "e& Pay integration · invoice engine live · task engine + service desk live."),
    (9,  "People Agent (WPS, payroll)",                       "Wave 4", "W25",  573000, "MoHRE/WPS integration · payroll engine live · onboarding workflow accepted."),
    (10, "Hardening, audit, handoff",                         "Wave 4", "W30",  573000, "External pen-test passed · NESA/TDRA controls audited · platform handoff accepted."),
]

for i, (idx, label, wave, week, amount, criteria) in enumerate(milestones):
    r = hdr_r + 1 + i
    ws.cell(row=r, column=2, value=idx).font = F_LBL; ws.cell(row=r, column=2).alignment = ALIGN_C; ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=3, value=label).font = F_LBL; ws.cell(row=r, column=3).alignment = ALIGN_L; ws.cell(row=r, column=3).border = BORDER
    wc = ws.cell(row=r, column=4, value=wave); wc.font = F_TXT; wc.alignment = ALIGN_C; wc.border = BORDER
    tc = ws.cell(row=r, column=5, value=week); tc.font = F_TXT; tc.alignment = ALIGN_C; tc.border = BORDER
    ac = ws.cell(row=r, column=6, value=amount); style_total(ac); ac.number_format = "$#,##0"
    nc = ws.cell(row=r, column=7, value=criteria); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

sum_r = hdr_r + 1 + len(milestones)
ws.cell(row=sum_r, column=3, value="TOTAL programme envelope").font = F_TOT
ws.cell(row=sum_r, column=3).fill = FILL_TOTAL; ws.cell(row=sum_r, column=3).border = BORDER; ws.cell(row=sum_r, column=3).alignment = ALIGN_R
amt_sum = ws.cell(row=sum_r, column=6, value=f"=SUM(F{hdr_r + 1}:F{sum_r - 1})")
style_total(amt_sum); amt_sum.number_format = "$#,##0"
amt_sum.font = Font(name="Calibri", size=11, bold=True, color=NAVY)

# Cash flow chart (months from kickoff)
cfh_row = sum_r + 3
ws.cell(row=cfh_row, column=2, value="Indicative cash flow — months from kickoff").font = F_H1

months = ["M0 (kickoff)", "M1", "M2", "M3", "M4", "M5", "M6", "M7"]   # 30 weeks ≈ 7.5 months
hdr_r2 = cfh_row + 1
ws.cell(row=hdr_r2, column=2, value="Item").font = F_HDR
ws.cell(row=hdr_r2, column=2).fill = FILL_HDR; ws.cell(row=hdr_r2, column=2).alignment = ALIGN_C; ws.cell(row=hdr_r2, column=2).border = BORDER
for j, m in enumerate(months):
    c = ws.cell(row=hdr_r2, column=3 + j, value=m)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Cash IN by month: M0=W1 (M1 milestone), M1=W4 (M2), M2=W8 (M3), M3=W12 (M4), M3.5=W13 → put M3 (M5 + M4 align), simplify mapping:
#   M0 = M1@W1 = $287k
#   M1 = M2@W4 = $287k
#   M2 = M3@W8 = $287k
#   M3 = M4@W12 + M5@W13 = $287k + $287.5k
#   M4 = M6@W18 + M7@W19 = $287.5k + $287.5k
#   M5 = M8@W24 + M9@W25 = $287.5k + $573k
#   M6 = (gap)
#   M7 = M10@W30 = $573k
in_row = hdr_r2 + 1
ws.cell(row=in_row, column=2, value="Cash IN (payments from e&)").font = F_LBL
ws.cell(row=in_row, column=2).alignment = ALIGN_L; ws.cell(row=in_row, column=2).border = BORDER

cash_in_per_month = [
    "=F6",                  # M0 (M1 milestone, row 6 of Sheet 9)
    "=F7",                  # M1 (M2)
    "=F8",                  # M2 (M3)
    "=F9+F10",              # M3 (M4 + M5)
    "=F11+F12",             # M4 (M6 + M7)
    "=F13+F14",             # M5 (M8 + M9)
    0,                       # M6
    "=F15",                  # M7 (M10)
]
for j, v in enumerate(cash_in_per_month):
    cell = ws.cell(row=in_row, column=3 + j, value=v)
    style_formula(cell); cell.number_format = "$#,##0"

# Cumulative
cum_row = in_row + 1
ws.cell(row=cum_row, column=2, value="Cumulative cash IN").font = F_TOT
ws.cell(row=cum_row, column=2).fill = FILL_TOTAL; ws.cell(row=cum_row, column=2).border = BORDER
ws.cell(row=cum_row, column=3, value=f"=C{in_row}")
style_total(ws.cell(row=cum_row, column=3)); ws.cell(row=cum_row, column=3).number_format = "$#,##0"
for j in range(1, len(months)):
    prev = get_column_letter(2 + j)
    cur = get_column_letter(3 + j)
    cell = ws.cell(row=cum_row, column=3 + j, value=f"={prev}{cum_row}+{cur}{in_row}")
    style_total(cell); cell.number_format = "$#,##0"

ws.cell(row=cum_row + 2, column=2, value="Note: cash-in maps milestone weeks to month buckets. Out-flow / cost burn modelled on Sheet 4 phase totals.").font = F_NOTE


# =============================================================================
# SHEET 10 — Sensitivity
# =============================================================================
ws = wb.create_sheet("10. Sensitivity")
set_widths(ws, {"A": 3, "B": 38, "C": 16, "D": 16, "E": 16, "F": 18, "G": 50})

ws["B2"] = "What-If Scenarios — How Much Do We Earn (3-year USD)"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: What if fewer SMBs sign up? What if ARPU is lower? This grid stress-tests our 3-year revenue under 9 scenarios."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

# A. Adoption scenarios — EOY customers Y1/Y2/Y3
ws["B5"] = "A. Adoption scenarios — EOY customers per year"
ws["B5"].font = F_H1

hdr_r = 6
for i, h in enumerate(["Scenario", "Y1 EOY", "Y2 EOY", "Y3 EOY", "Notes"]):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

scenarios = [
    ("Low — slow ramp",          [12000, 30000, 50000], "Conservative: half of agreement."),
    ("Mid — agreement baseline", [24000, 62000, 85000], "Per FINAL agreement."),
    ("High — strong adoption",   [40000, 95000, 130000], "If e& sales engine fires; +60%."),
]
for i, (label, vals, note) in enumerate(scenarios):
    r = hdr_r + 1 + i
    ws.cell(row=r, column=2, value=label).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    for j, v in enumerate(vals):
        cell = ws.cell(row=r, column=3 + j, value=v)
        style_input(cell); cell.number_format = "#,##0"
    nc = ws.cell(row=r, column=6, value=note); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# B. ARPU x rev-share matrix
arpu_label_row = hdr_r + 1 + len(scenarios) + 2
ws.cell(row=arpu_label_row, column=2, value="B. AIdeology 3-year revenue (USD) — adoption × ARPU (declining 35/35/28% rev-share applied)").font = F_H1

# Build a matrix: scenarios × ARPU levels (low/mid/high in AED)
mhdr_r = arpu_label_row + 1
arpu_levels = [(220, "Low"), (350, "Mid (per agt)"), (450, "High")]
mhdrs = ["Adoption × ARPU"] + [f"ARPU AED {a} ({lbl})" for a, lbl in arpu_levels] + ["", "Notes"]
for i, h in enumerate(mhdrs):
    c = ws.cell(row=mhdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Calc per scenario:
#   avg_yi = (eoy_{i-1} + eoy_i) / 2  (with eoy_0 = 0)
#   gross_yi = avg_yi * 12 * arpu
#   our_yi = gross_yi * pct_yi   (Y1=35%, Y2=35%, Y3=28%)
#   3yr USD = SUM(our_y1..3) / FX
FX_REF = "'1. Cover & Assumptions'!$C$9"

for i, (label, eoy_vals, _note) in enumerate(scenarios):
    r = mhdr_r + 1 + i
    sc_row = hdr_r + 1 + i      # row in adoption section
    ws.cell(row=r, column=2, value=label).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    for j, (arpu, lbl) in enumerate(arpu_levels):
        # Build formula referencing scenario eoy values + arpu constant
        col = 3 + j
        # avg_y1 = C{sc_row}/2; avg_y2 = (C{sc_row}+D{sc_row})/2; avg_y3 = (D{sc_row}+E{sc_row})/2
        # rev_y1 = avg_y1*12*arpu*0.35; rev_y2 = avg_y2*12*arpu*0.35; rev_y3 = avg_y3*12*arpu*0.28
        formula = (
            f"=( (C{sc_row}/2)*12*{arpu}*0.35"
            f" + ((C{sc_row}+D{sc_row})/2)*12*{arpu}*0.35"
            f" + ((D{sc_row}+E{sc_row})/2)*12*{arpu}*0.28"
            f" ) / {FX_REF}"
        )
        cell = ws.cell(row=r, column=col, value=formula)
        style_total(cell); cell.number_format = "$#,##0"
    nc = ws.cell(row=r, column=6, value="3-year USD rev-share at declining 35/35/28%.")
    nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# C. Total AIdeology 3-year revenue including fixed fees envelope
total_label_row = mhdr_r + 1 + len(scenarios) + 2
ws.cell(row=total_label_row, column=2, value="C. Total AIdeology 3-year revenue (USD) — fixed fees + rev-share").font = F_H1

thdr_r = total_label_row + 1
mhdrs2 = ["Adoption × ARPU"] + [f"ARPU AED {a} ({lbl})" for a, lbl in arpu_levels] + ["", "Notes"]
for i, h in enumerate(mhdrs2):
    c = ws.cell(row=thdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

ENVELOPE_REF = "'7. Pricing Summary'!$C$13"   # row 13 of Sheet 7 (TOTAL row)

for i, (label, _eoy_vals, _note) in enumerate(scenarios):
    r = thdr_r + 1 + i
    rs_row = mhdr_r + 1 + i
    ws.cell(row=r, column=2, value=label).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    for j in range(len(arpu_levels)):
        col_letter = get_column_letter(3 + j)
        cell = ws.cell(row=r, column=3 + j, value=f"={ENVELOPE_REF}+{col_letter}{rs_row}")
        style_total(cell); cell.number_format = "$#,##0"
    nc = ws.cell(row=r, column=6, value="Fixed-fee envelope (Sheet 7) + 3-year rev-share at scenario.")
    nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER
# =============================================================================
# SHEET 11 — Acquisition & Valuation Pathway
# =============================================================================
ws = wb.create_sheet("11. Acquisition & Valuation")
set_widths(ws, {"A": 3, "B": 38, "C": 22, "D": 22, "E": 22, "F": 22, "G": 50})

ws["B2"] = "Acquisition & Valuation Pathway — The Bigger Prize"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: by Year 3-4, e& is paying us 20-28% of revenue forever. Buying us out becomes cheaper. This sheet shows what we're worth."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

# A. Year 4-5 run-rate forecast (using FINAL agreement assumptions)
ws["B5"] = "A. Year 4-5 run rates (steady-state SaaS economics)"
ws["B5"].font = F_H1

hdr_r = 6
hdrs = ["Metric", "Year 1", "Year 2", "Year 3", "Year 4 (proj.)", "Notes"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Row 7 - customers EOY
ws.cell(row=7, column=2, value="Customers EOY").font = F_LBL; ws.cell(row=7, column=2).border = BORDER
for j, src in enumerate(["C", "D", "E"]):
    cell = ws.cell(row=7, column=3 + j, value=f"='8. Revenue Share Model'!{src}7")
    style_formula(cell); cell.number_format = "#,##0"
# Y4 projection: linear growth from Y3 EOY (~+15k)
y4 = ws.cell(row=7, column=6, value="=E7+15000"); style_input(y4); y4.number_format = "#,##0"

# Row 8 - ARPU AED
ws.cell(row=8, column=2, value="ARPU (AED/mo)").font = F_LBL; ws.cell(row=8, column=2).border = BORDER
for j, src in enumerate(["C", "D", "E"]):
    cell = ws.cell(row=8, column=3 + j, value=f"='8. Revenue Share Model'!{src}9")
    style_formula(cell); cell.number_format = "#,##0"
y4 = ws.cell(row=8, column=6, value="=E8+30"); style_input(y4); y4.number_format = "#,##0"

# Row 9 - Gross AED
ws.cell(row=9, column=2, value="e& gross SaaS rev (AED)").font = F_LBL; ws.cell(row=9, column=2).border = BORDER
for j in range(3):
    col = 3 + j; col_letter = get_column_letter(col)
    cell = ws.cell(row=9, column=col, value=f"='8. Revenue Share Model'!{col_letter}11")
    style_formula(cell); cell.number_format = "#,##0,,\"M AED\""
# Y4 projection
y4 = ws.cell(row=9, column=6, value="=F7*F8*12"); style_total(y4); y4.number_format = "#,##0,,\"M AED\""

# Row 10 - rev-share %
ws.cell(row=10, column=2, value="AIdeology rev-share %").font = F_LBL; ws.cell(row=10, column=2).border = BORDER
for j, src in enumerate(["C", "D", "E"]):
    cell = ws.cell(row=10, column=3 + j, value=f"='8. Revenue Share Model'!{src}12")
    style_formula(cell); cell.number_format = "0%"
y4 = ws.cell(row=10, column=6, value="='1. Cover & Assumptions'!$C$28"); style_formula(y4); y4.number_format = "0%"

# Row 11 - AIdeology AED
ws.cell(row=11, column=2, value="AIdeology rev-share (AED)").font = F_TOT
ws.cell(row=11, column=2).fill = FILL_TOTAL; ws.cell(row=11, column=2).border = BORDER
for j in range(4):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=11, column=3 + j, value=f"={col_letter}9*{col_letter}10")
    style_total(cell); cell.number_format = "#,##0,,\"M AED\""

# Row 12 - AIdeology USD
ws.cell(row=12, column=2, value="AIdeology rev-share (USD)").font = F_TOT
ws.cell(row=12, column=2).fill = FILL_TOTAL; ws.cell(row=12, column=2).border = BORDER
for j in range(4):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=12, column=3 + j, value=f"={col_letter}11/'1. Cover & Assumptions'!$C$9")
    style_total(cell); cell.number_format = "$#,##0"

# B. Valuation scenarios (8x / 10x / 15x EBITDA proxy on Y4 run rate)
ws.cell(row=14, column=2, value="B. Valuation scenarios (Year 4 run-rate × SaaS multiple)").font = F_H1

hdr_v = 15
hvs = ["Multiple basis", "Conservative (8× EBITDA)", "Base case (10× plus regional licence)", "Optimistic (15× regional moat)", "Notes"]
for i, h in enumerate(hvs):
    c = ws.cell(row=hdr_v, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Row 16 — AED valuation
ws.cell(row=16, column=2, value="AIdeology platform valuation (AED)").font = F_TOT
ws.cell(row=16, column=2).fill = FILL_GOLD; ws.cell(row=16, column=2).border = BORDER
# Y4 AIdeology AED rev (cell F11)
ws.cell(row=16, column=3, value="=F11*8");  style_total(ws.cell(row=16, column=3));  ws.cell(row=16, column=3).number_format = "#,##0,,\"M AED\""; ws.cell(row=16, column=3).fill = FILL_GOLD
ws.cell(row=16, column=4, value="=F11*10"); style_total(ws.cell(row=16, column=4)); ws.cell(row=16, column=4).number_format = "#,##0,,\"M AED\""; ws.cell(row=16, column=4).fill = FILL_GOLD
ws.cell(row=16, column=5, value="=F11*15"); style_total(ws.cell(row=16, column=5)); ws.cell(row=16, column=5).number_format = "#,##0,,\"M AED\""; ws.cell(row=16, column=5).fill = FILL_GOLD
ws.cell(row=16, column=6, value="Multiple × Y4 AIdeology rev-share. Aligns with ~$200-700M USD range.").font = F_NOTE
ws.cell(row=16, column=6).border = BORDER

# Row 17 — USD valuation
ws.cell(row=17, column=2, value="AIdeology platform valuation (USD)").font = F_TOT
ws.cell(row=17, column=2).fill = FILL_GOLD; ws.cell(row=17, column=2).border = BORDER
for j in range(3):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=17, column=3 + j, value=f"={col_letter}16/'1. Cover & Assumptions'!$C$9")
    style_total(cell); cell.number_format = "$#,##0"; cell.fill = FILL_GOLD

# C. Acquisition mechanics
ws.cell(row=19, column=2, value="C. Why e& is likely to acquire by Year 3-4").font = F_H1

mech_lines = [
    "1. Cumulative payments: by Y4, e& has paid us ~AED 173M cumulative rev-share. By Y5, ~AED 280M+. Buying us out for AED 1-3B caps the bleed.",
    "2. Strategic moat: platform IP we hold can license to Saudi Telecom / Maroc Telecom — e& wants exclusivity in MENA.",
    "3. Talent & continuity: our team becomes their team. They get the institutional knowledge instead of running a perpetual contract.",
    "4. Independence: e& AI team owns agents (after Wave 4). Owning the platform too completes their stack.",
    "5. Exit option for AIdeology: cash + equity rolled into e& makes founders / investors realise 10-30× return on seed.",
]
for i, line in enumerate(mech_lines):
    ws.merge_cells(start_row=20 + i, start_column=2, end_row=20 + i, end_column=6)
    c = ws.cell(row=20 + i, column=2, value=line); c.font = F_NOTE; c.alignment = ALIGN_L

# D. Acquisition timing & price band
ws.cell(row=27, column=2, value="D. Acquisition timing & price band").font = F_H1

hdr_d = 28
for i, h in enumerate(["Window", "Trigger", "Price band (USD)", "Strategic context", "Notes"]):
    c = ws.cell(row=hdr_d, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

acq_rows = [
    ("Year 2 (M18-24)",  "Wave 4 complete · 60k+ customers · proven UAE market", "$200-400M",   "Defensive — e& pre-empts.",            "Earliest realistic offer; full earnout."),
    ("Year 3 (M24-36)",  "Saudi/Morocco scaling · 85k+ customers · 28% rev-share active", "$400-800M",   "Most likely window.",            "e& AI team capable; rev-share cost compounds."),
    ("Year 4 (M36-48)",  "100k+ customers · multi-telco licensing real",         "$800M-1.5B",  "Strategic — e& wants exclusivity.",    "Other telcos (STC, Zain) in play."),
    ("Year 5+",          "Regional category leader · 5+ telcos / 500k+ customers", "$1.5-3B",     "Premium — competitive bidding.",       "IPO becomes alternative."),
]
for i, (window, trigger, band, ctx, note) in enumerate(acq_rows):
    r = hdr_d + 1 + i
    ws.cell(row=r, column=2, value=window).font = F_LBL; ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=3, value=trigger).font = F_TXT; ws.cell(row=r, column=3).alignment = ALIGN_L; ws.cell(row=r, column=3).border = BORDER
    bc = ws.cell(row=r, column=4, value=band); bc.font = F_VAL; bc.alignment = ALIGN_C; bc.border = BORDER; bc.fill = FILL_GOLD
    ws.cell(row=r, column=5, value=ctx).font = F_TXT; ws.cell(row=r, column=5).alignment = ALIGN_L; ws.cell(row=r, column=5).border = BORDER
    ws.cell(row=r, column=6, value=note).font = F_NOTE; ws.cell(row=r, column=6).alignment = ALIGN_L; ws.cell(row=r, column=6).border = BORDER


# =============================================================================
# Save
# =============================================================================
import os
os.makedirs("Internal Analysis", exist_ok=True)
wb.save(OUT)
print(f"✓ Saved: {OUT}")
