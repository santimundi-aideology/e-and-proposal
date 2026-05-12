"""
Build the ENTERPRISE internal pricing & staffing analysis workbook.

Mirrors the structure of the SMB workbook but scoped to the e& B2B
ENTERPRISE & GOVERNMENT pillar (Pillar 02 in the commercial framework).

Key principle the Enterprise model is built on:
    The SMB platform is the BASELINE. The Enterprise platform is a
    set of extensions on top (P1 enterprise-readiness + P2 government
    layer). Agents are ADAPTED from the SMB marketplace versions —
    we charge for the enterprise-specific delta (multi-branch, ERP/
    CRM/HIS connectors, compliance logging, fine-tuning, RBAC, audit
    trail) NOT for rebuilding the agent. That is why Tier 1 enterprise
    agents cost dramatically less than the original SMB Wave 2-4 builds.

Pillars covered:
    - Wave E0  — Enterprise platform extensions (P1 + P2 layers)
    - Tier 1   — Adapted SMB Agents (Customer / Sales / Comms / Finance / Ops / People)
                 with enterprise connectors and fine-tuning
    - Tier 2   — Custom AI Solutions (Contact Centre AI, Document
                 Intelligence, Approval Orchestration, Predictive
                 Maintenance, Security & Compliance)
    - Tier 3   — Sovereign & On-Prem deployments (customer site,
                 HPC reference architectures held in hand for vendor
                 BoM agility — Cisco / Dell / HPE / Lenovo / Supermicro)
    - Custom   — Bespoke AIdeology engineering for unique needs
    - Managed  — Recurring monthly managed service per client

Deployment options modelled per Tier 2 / Tier 3:
    - e& cloud (lowest deployment uplift)
    - G42 / Core42 sovereign cloud
    - On-premise (highest uplift; uses pre-prepared HPC RA BoMs)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "Internal Analysis/AIdeology_eand_Enterprise_Pricing_Analysis_v0.1.xlsx"

# -----------------------------------------------------------------------------
# Style palette (matches SMB / ADNOC workbooks)
# -----------------------------------------------------------------------------
NAVY = "FF1F3A5F"
WHITE = "FFFFFFFF"
GREY = "FF808080"
INK = "FF000000"
BLUE = "FF0000FF"
GREEN = "FF008000"
RED = "FFC8102E"
PURPLE = "FF7A52F4"             # Enterprise accent (Pillar 02 violet)
LIGHT_YELLOW = "FFFFF2CC"
LIGHT_GREY = "FFF5F8FB"
LIGHT_GREEN = "FFE2F0D9"
LIGHT_BLUE = "FFDDEBF7"
LIGHT_VIOLET = "FFEEEAFE"

thin = Side(border_style="thin", color="FFD0D5DC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

F_TITLE = Font(name="Calibri", size=16, bold=True, color=NAVY)
F_SUB   = Font(name="Calibri", size=11, color=GREY, italic=True)
F_NOTE  = Font(name="Calibri", size=10, color=GREY)
F_H1    = Font(name="Calibri", size=12, bold=True, color="FF2E75B6")
F_HDR   = Font(name="Calibri", size=10, bold=True, color=WHITE)
F_LBL   = Font(name="Calibri", size=10, bold=True, color=INK)
F_TXT   = Font(name="Calibri", size=10, color=INK)
F_INPUT = Font(name="Calibri", size=10, bold=False, color=BLUE)
F_FORMULA = Font(name="Calibri", size=10, color=GREEN)
F_TOT   = Font(name="Calibri", size=10, bold=True, color=INK)

FILL_HDR    = PatternFill("solid", fgColor=NAVY)
FILL_HDR_V  = PatternFill("solid", fgColor=PURPLE)
FILL_INPUT  = PatternFill("solid", fgColor=LIGHT_YELLOW)
FILL_DATA   = PatternFill("solid", fgColor=LIGHT_GREY)
FILL_TOTAL  = PatternFill("solid", fgColor=LIGHT_GREEN)
FILL_SUB    = PatternFill("solid", fgColor=LIGHT_BLUE)
FILL_VIOLET = PatternFill("solid", fgColor=LIGHT_VIOLET)

ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center", wrap_text=True)

F_BIG = Font(name="Calibri", size=24, bold=True, color=NAVY)
F_MID = Font(name="Calibri", size=14, bold=True, color=NAVY)
F_DESC = Font(name="Calibri", size=11, color="FF444444")
F_VAL_BIG = Font(name="Calibri", size=14, bold=True, color=NAVY)
F_VAL = Font(name="Calibri", size=11, bold=True, color=INK)
F_LABEL = Font(name="Calibri", size=11, color=GREY)
FILL_ACCENT = PatternFill("solid", fgColor="FFE8F0FE")
FILL_WARN = PatternFill("solid", fgColor="FFFFF3E0")
FILL_GREEN_LT = PatternFill("solid", fgColor="FFE8F5E9")


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def style_input(cell):
    cell.font = F_INPUT
    cell.fill = FILL_INPUT
    cell.border = BORDER
    cell.alignment = ALIGN_R

def style_formula(cell):
    cell.font = F_FORMULA
    cell.fill = FILL_DATA
    cell.border = BORDER
    cell.alignment = ALIGN_R

def style_total(cell):
    cell.font = F_TOT
    cell.fill = FILL_TOTAL
    cell.border = BORDER
    cell.alignment = ALIGN_R

def style_label(cell):
    cell.font = F_LBL
    cell.alignment = ALIGN_L
    cell.border = BORDER

def style_text(cell):
    cell.font = F_TXT
    cell.alignment = ALIGN_L
    cell.border = BORDER

def set_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def violet_header(ws, row, col_start, col_end, text):
    """Section header bar in violet to mark Enterprise sections."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    ws.cell(row=row, column=col_start, value=text).font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    ws.cell(row=row, column=col_start).fill = FILL_HDR_V
    ws.cell(row=row, column=col_start).alignment = ALIGN_L
    ws.cell(row=row, column=col_start).border = BORDER
    for c in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=c).fill = FILL_HDR_V
        ws.cell(row=row, column=c).border = BORDER

def summary_box(ws, row, col, label, formula, fmt="$#,##0", width=2):
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = F_LABEL; lc.alignment = ALIGN_L; lc.border = BORDER
    lc.fill = FILL_VIOLET
    vc = ws.cell(row=row + 1, column=col, value=formula)
    vc.font = F_VAL_BIG; vc.alignment = ALIGN_L; vc.border = BORDER
    vc.number_format = fmt; vc.fill = FILL_VIOLET
    if width > 1:
        for c in range(col + 1, col + width):
            ws.cell(row=row, column=c).fill = FILL_VIOLET
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row + 1, column=c).fill = FILL_VIOLET
            ws.cell(row=row + 1, column=c).border = BORDER


# =============================================================================
wb = Workbook()
wb.remove(wb.active)


# =============================================================================
# SHEET 0 — Executive Summary
# =============================================================================
ws = wb.create_sheet("0. Executive Summary")
set_widths(ws, {"A": 3, "B": 30, "C": 22, "D": 22, "E": 22, "F": 22, "G": 22})

ws.merge_cells("B2:G2")
ws["B2"] = "AIdeology × e&  —  ENTERPRISE Internal Pricing Summary"
ws["B2"].font = F_BIG; ws["B2"].alignment = ALIGN_L

ws.merge_cells("B3:G3")
ws["B3"] = "Pillar 02 (Enterprise & Government). One page. Detail in tabs 1-11."
ws["B3"].font = F_DESC; ws["B3"].alignment = ALIGN_L

# A. WHAT ARE WE BUILDING?
violet_header(ws, 5, 2, 7, "A.  WHAT ARE WE BUILDING?")
desc_lines = [
    "We extend the SMB platform with an enterprise-readiness layer (RBAC, audit, trust tiers, ERP/CRM/HIS integration, fine-tuning) and",
    "deliver three engagement tiers to e&'s enterprise & government customers:",
    "  Tier 1  — adapted SMB agents with enterprise connectors (deeper, multi-branch, compliance, fine-tuned)",
    "  Tier 2  — custom AI solutions including AI Contact Centre, Document Intelligence, Approvals, Predictive Maint., Security",
    "  Tier 3  — sovereign / on-premise deployments (customer site) using pre-prepared HPC RA BoMs for vendor agility",
    "Each project can deploy on G42 / Core42 sovereign cloud, e& cloud, or customer on-prem. Custom bespoke engineering on demand.",
]
for i, line in enumerate(desc_lines):
    ws.merge_cells(start_row=6+i, start_column=2, end_row=6+i, end_column=7)
    c = ws.cell(row=6+i, column=2, value=line)
    c.font = F_DESC; c.alignment = ALIGN_L

# B. WHAT DOES IT COST US?
r = 13
violet_header(ws, r, 2, 7, "B.  WHAT DOES IT COST US?  (per build / per project — internal)")
summary_box(ws, r+1, 2, "Wave E0 — Platform extensions (one-off)", "='8. Pricing Summary'!C9")
summary_box(ws, r+1, 4, "Tier 1 cost — average per adapted agent", "='8. Pricing Summary'!C20")
summary_box(ws, r+1, 6, "Tier 2 cost — average per custom solution", "='8. Pricing Summary'!C30")

r2 = r + 4
summary_box(ws, r2, 2, "Tier 3 cost — average sovereign / on-prem project", "='8. Pricing Summary'!C36")
summary_box(ws, r2, 4, "Custom bespoke — average ~22-week build", "='7. Custom Bespoke'!C26")
summary_box(ws, r2, 6, "Managed service — monthly cost per client", "='8. Pricing Summary'!C41")

# C. WHAT DO WE CHARGE e&?
r = 21
violet_header(ws, r, 2, 7, "C.  WHAT DO WE CHARGE e&?  (price to e&)")
summary_box(ws, r+1, 2, "Wave E0 — Platform extensions price", "='8. Pricing Summary'!E9")
summary_box(ws, r+1, 4, "Tier 1 — average price per adapted agent", "='8. Pricing Summary'!E20")
summary_box(ws, r+1, 6, "Tier 2 — average price per custom solution", "='8. Pricing Summary'!E30")

r2 = r + 4
summary_box(ws, r2, 2, "Tier 3 — average price per sovereign project", "='8. Pricing Summary'!E36")
summary_box(ws, r2, 4, "Custom bespoke — average price", "='7. Custom Bespoke'!E26")
summary_box(ws, r2, 6, "Managed service — monthly price per client", "='8. Pricing Summary'!E41")

# D. 3-YEAR REVENUE FORECAST
r = 29
violet_header(ws, r, 2, 7, "D.  3-YEAR REVENUE FORECAST  (mid scenario)")

th_row = r + 1
for i, h in enumerate(["", "Year 1", "Year 2", "Year 3", "3-Year Total"]):
    c = ws.cell(row=th_row, column=2+i, value=h)
    c.font = F_LBL; c.fill = FILL_VIOLET; c.alignment = ALIGN_C; c.border = BORDER

forecast_items = [
    ("# Projects (Y1: 4, Y2: 11, Y3: 18 mid-scenario)", "='9. Revenue Forecast'!C9", "='9. Revenue Forecast'!D9", "='9. Revenue Forecast'!E9", None, "#,##0"),
    ("Gross project & build revenue (e& sale price)", "='9. Revenue Forecast'!C19", "='9. Revenue Forecast'!D19", "='9. Revenue Forecast'!E19", None, "$#,##0"),
    ("Managed service recurring revenue", "='9. Revenue Forecast'!C24", "='9. Revenue Forecast'!D24", "='9. Revenue Forecast'!E24", None, "$#,##0"),
    ("AIdeology share (60% of project + 60% managed)", "='9. Revenue Forecast'!C30", "='9. Revenue Forecast'!D30", "='9. Revenue Forecast'!E30", None, "$#,##0"),
    ("e& share (40% project + 40% managed + hosting)", "='9. Revenue Forecast'!C31", "='9. Revenue Forecast'!D31", "='9. Revenue Forecast'!E31", None, "$#,##0"),
]
for i, (label, y1, y2, y3, _, fmt) in enumerate(forecast_items):
    row = th_row + 1 + i
    ws.cell(row=row, column=2, value=label).font = F_LBL; ws.cell(row=row, column=2).border = BORDER
    for j, v in enumerate([y1, y2, y3]):
        cell = ws.cell(row=row, column=3+j, value=v)
        cell.font = F_VAL; cell.number_format = fmt; cell.border = BORDER
    tot = ws.cell(row=row, column=6, value=f"=SUM(C{row}:E{row})")
    tot.font = F_VAL; tot.number_format = fmt; tot.border = BORDER
    if "AIdeology share" in label:
        for j in range(4):
            ws.cell(row=row, column=3+j).fill = FILL_GREEN_LT
            ws.cell(row=row, column=3+j).font = F_VAL_BIG

# E. THREE PRICING POSTURES
r = 37
violet_header(ws, r, 2, 7, "E.  THREE PRICING POSTURES")
r2 = r + 1
for i, h in enumerate(["", "Conservative", "Mid (baseline)", "Aggressive"]):
    c = ws.cell(row=r2, column=2+i, value=h)
    c.font = F_LBL; c.fill = FILL_VIOLET; c.alignment = ALIGN_C; c.border = BORDER

opt_data = [
    ("Tier 1 price per agent", 30000, 55000, 80000),
    ("Tier 2 price per project", 300000, 450000, 600000),
    ("Tier 3 price per project", 600000, 800000, 1000000),
    ("Best for", "AIdeology cash certainty", "JSX framework baseline", "Maximise margin on flagships"),
    ("Risk", "Leaves money on the table", "Balanced", "Slower pipeline"),
]
for i, (label, *vals) in enumerate(opt_data):
    row = r2 + 1 + i
    ws.cell(row=row, column=2, value=label).font = F_LBL; ws.cell(row=row, column=2).border = BORDER
    for j, v in enumerate(vals):
        cell = ws.cell(row=row, column=3+j, value=v); cell.border = BORDER
        if isinstance(v, (int, float)):
            cell.font = F_VAL; cell.number_format = "$#,##0"
        else:
            cell.font = F_DESC; cell.alignment = ALIGN_C

# F. KEY INSIGHTS
r = 44
violet_header(ws, r, 2, 7, "F.  KEY INSIGHTS")
ws.merge_cells(f"B{r+1}:G{r+1}")
ws.cell(row=r+1, column=2, value="SMB platform is the baseline. Tier 1 enterprise agents are NOT rebuilt — we add multi-branch, deep ERP/CRM/HIS connectors, compliance logging, and fine-tuning ON TOP. Average Tier 1 internal cost ~$100K (vs ~$240K to build an SMB wave from zero). For repeat agents on the same customer, the ERP integration work is amortised → second & third agents cost ~50% less.").font = Font(name="Calibri", size=11, bold=True, color="FF1B5E20")
ws.cell(row=r+1, column=2).fill = FILL_GREEN_LT; ws.cell(row=r+1, column=2).alignment = ALIGN_L; ws.cell(row=r+1, column=2).border = BORDER
for c in range(3, 8): ws.cell(row=r+1, column=c).fill = FILL_GREEN_LT; ws.cell(row=r+1, column=c).border = BORDER

ws.merge_cells(f"B{r+2}:G{r+2}")
ws.cell(row=r+2, column=2, value="On-premise deployments are the most expensive (Tier 3) — but with pre-prepared HPC reference architectures (Cisco / Dell / HPE / Lenovo / Supermicro NVIDIA-Certified BoMs), AIdeology can spin up a customer-specific BoM in days, not weeks. See Sheet 10 for the BoM library.").font = F_DESC
ws.cell(row=r+2, column=2).fill = FILL_VIOLET; ws.cell(row=r+2, column=2).alignment = ALIGN_L; ws.cell(row=r+2, column=2).border = BORDER
for c in range(3, 8): ws.cell(row=r+2, column=c).fill = FILL_VIOLET; ws.cell(row=r+2, column=c).border = BORDER

ws.merge_cells(f"B{r+3}:G{r+3}")
ws.cell(row=r+3, column=2, value="AI Contact Centre Software for Enterprises is delivered as the headline Tier 2 custom solution (Sheet 5). It packages voice + WhatsApp + web + agent assist + supervisor analytics on the AIdeology platform with deep CRM/ERP/HIS integration. Internal cost ~$374K → price to e& ~$680K @ 45% margin (above JSX $300-600K — verify per deal).").font = F_DESC
ws.cell(row=r+3, column=2).fill = FILL_ACCENT; ws.cell(row=r+3, column=2).alignment = ALIGN_L; ws.cell(row=r+3, column=2).border = BORDER
for c in range(3, 8): ws.cell(row=r+3, column=c).fill = FILL_ACCENT; ws.cell(row=r+3, column=c).border = BORDER

ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1


# =============================================================================
# SHEET 1 — Cover & Assumptions
# =============================================================================
ws = wb.create_sheet("1. Cover & Assumptions")
set_widths(ws, {"A": 3, "B": 46, "C": 16, "D": 10, "E": 64})

ws["B2"] = "AIdeology — Enterprise Internal Pricing & Staffing Analysis"
ws["B2"].font = F_TITLE
ws["B3"] = "e& B2B Acceleration · Pillar 02 — Enterprise & Government · v0.1 DRAFT"
ws["B3"].font = F_SUB
ws["B4"] = "INTERNAL — NOT FOR e&. Pricing model for AIdeology team only."
ws["B4"].font = Font(name="Calibri", size=10, color=RED, italic=True)
ws["B5"] = "Date: 2026-05-10"
ws["B5"].font = F_NOTE

ws["B7"] = "How to use this workbook"
ws["B7"].font = F_H1
how_to = [
    "• Blue-text / yellow-shaded cells are editable inputs (rates, p-weeks, margins, deal counts).",
    "• Green-text cells are formulas — do not overwrite.",
    "• Sheet 2: Roles & Rates — same SMB roles + 4 enterprise-specific roles (ERP architect, Compliance SME, Industry SME, Fine-Tuning Eng).",
    "• Sheet 3: Wave E0 — Platform extensions on top of SMB platform (P1 + P2 layers, ~12 weeks).",
    "• Sheet 4: Tier 1 — Adapted SMB Agents (delta cost only — agents reuse SMB build).",
    "• Sheet 5: Tier 2 — Custom AI Solutions (Contact Centre, Doc Intel, Approvals, Pred Maint, Security).",
    "• Sheet 6: Tier 3 — Sovereign / On-Prem (uses pre-prepared HPC RA BoMs).",
    "• Sheet 7: Custom Bespoke — Unique builds priced T&M or fixed bid.",
    "• Sheet 8: Pricing Summary — Cost + margin = price per tier.",
    "• Sheet 9: 3-Year Revenue Forecast — Pipeline × tier mix → AIdeology vs e& share.",
    "• Sheet 10: HPC RA Library — Pre-prepared BoMs from Cisco / Dell / HPE / Lenovo / Supermicro.",
    "• Sheet 11: Sensitivity — Pipeline volume × tier mix × revenue split scenarios.",
]
for i, line in enumerate(how_to):
    ws.cell(row=8 + i, column=2, value=line).font = F_NOTE

# Global assumptions
ws["B22"] = "Global assumptions"
ws["B22"].font = F_H1

assum_hdr = ["Assumption", "Value", "Unit", "Notes / source"]
for i, h in enumerate(assum_hdr):
    c = ws.cell(row=23, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

assum = [
    ("Working days per week", 5, "days", "Standard."),
    ("Hours per working day", 8, "hours", "Standard."),
    ("Currency", "USD", "—", "Internal model; price to e& issuable in USD or AED."),
    ("FX assumption (USD→AED)", 3.67, "rate", "Pegged."),
    ("Wave E0 — Platform extensions duration", 12, "weeks", "P1 enterprise-readiness + P2 government layer (RBAC, audit, trust tiers, ERP framework, fine-tuning sub-platform)."),
    ("Wave E0 target gross margin", 0.45, "%", "Reusable platform IP; reused across every Tier 1/2/3 deal."),
    ("Tier 1 target gross margin", 0.50, "%", "Adapted SMB agent — minimal new build, mostly configuration & connectors. High margin justified."),
    ("Tier 2 target gross margin", 0.45, "%", "Custom AI solution build — moderate complexity, controlled cloud."),
    ("Tier 3 target gross margin", 0.40, "%", "Sovereign / on-prem — highest delivery risk + customer-specific work."),
    ("Custom bespoke target gross margin", 0.50, "%", "Unique IP creation — premium pricing, risk-adjusted."),
    ("Managed service target gross margin", 0.55, "%", "Recurring monthly support; high margin once stable."),
    ("Contingency on direct cost (Tier 1)", 0.10, "%", "Buffer for unknowns."),
    ("Contingency on direct cost (Tier 2)", 0.12, "%", "Higher — custom build risk."),
    ("Contingency on direct cost (Tier 3)", 0.15, "%", "Highest — customer-site, networking, security review uncertainty."),
    ("Risk reserve (held internally)", 0.05, "%", "Held at AIdeology level — not passed to e& in price."),
    ("Revenue split — AIdeology share of project fee", 0.60, "%", "JSX framework: 60/40 AIdeology/e&. e& sells, owns customer & hosting."),
    ("Revenue split — e& share of project fee", 0.40, "%", "Mirror — must equal 1 - AIdeology share."),
    ("Managed service split — AIdeology share", 0.60, "%", "Same as project split for simplicity."),
    ("# Projects Y1 (mid scenario)", 4, "deals", "Per JSX: 3-8 deals Y1 → mid 4."),
    ("# Projects Y2 (mid scenario)", 11, "deals", "Per JSX: 8-15 deals Y2 → mid 11."),
    ("# Projects Y3 (mid scenario)", 18, "deals", "Per JSX: 15-25 deals Y3 → mid 18."),
    ("Tier mix — % Tier 1 (adapted agents)", 0.35, "%", "Smaller deals dominate count (multiple agents per logo)."),
    ("Tier mix — % Tier 2 (custom AI)", 0.45, "%", "Sweet spot — Contact Centre, Doc Intel etc."),
    ("Tier mix — % Tier 3 (sovereign / on-prem)", 0.15, "%", "Flagships — gov, banks, healthcare regulator."),
    ("Tier mix — % Custom bespoke", 0.05, "%", "Rare — unique IP requests."),
    ("Avg # Tier 1 agents per deal", 2.5, "agents", "Most enterprise customers buy a bundle of 2-4 adapted agents."),
    ("Managed service ARPU (mid)", 12000, "$/mo/client", "JSX range $5K-$25K/mo. Mid $12K."),
    ("Managed service uptake — % of project clients", 0.70, "%", "70% of project clients sign managed service after go-live."),
    ("Managed service ramp delay (months from project close)", 1, "months", "Begins immediately after go-live."),
]
for i, (label, value, unit, note) in enumerate(assum):
    r = 24 + i
    ws.cell(row=r, column=2, value=label).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    vc = ws.cell(row=r, column=3, value=value)
    style_input(vc)
    if isinstance(value, float) and value < 1 and label.endswith(("%",)) is False and "Tier mix" in label:
        vc.number_format = "0.0%"
    elif isinstance(value, float) and value < 1:
        vc.number_format = "0.0%"
    elif isinstance(value, float):
        vc.number_format = "#,##0.00"
    elif isinstance(value, int):
        vc.number_format = "#,##0"
    uc = ws.cell(row=r, column=4, value=unit); uc.font = F_TXT; uc.alignment = ALIGN_C; uc.border = BORDER
    nc = ws.cell(row=r, column=5, value=note); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# Disclaimers
disc_row = 24 + len(assum) + 2
ws.cell(row=disc_row, column=2, value="Important disclaimers").font = F_H1
disclaimers = [
    "• Daily rates in Sheet 2 are PLACEHOLDER values to be validated against AIdeology's actual cost structure.",
    "• Per-tier p-weeks in Sheets 3-7 are first-pass sizing. Adjust based on deal scope per opportunity.",
    "• Tier 1 enterprise agents reuse the SMB platform build — only the delta work is costed here.",
    "• Tier 3 BoM costs in Sheet 10 are vendor MSRP indicative; e& negotiating power expected to discount 20-35%.",
    "• Help AG security wrap costs (~$50K-$100K) included on Tier 3; verify with Help AG quote per project.",
    "• Cloud costs depend on environment (e& OCI / G42 Core42 / Azure CSP via e&); revise once landing zone confirmed per deal.",
    "• Voice/STT/TTS costs (Deepgram, ElevenLabs, Azure Speech) estimated; replace with vendor quotes for large CCaaS deals.",
]
for i, line in enumerate(disclaimers):
    ws.cell(row=disc_row + 1 + i, column=2, value=line).font = F_NOTE


# =============================================================================
# SHEET 2 — Roles & Rates
# =============================================================================
ws = wb.create_sheet("2. Roles & Rates")
set_widths(ws, {"A": 3, "B": 46, "C": 14, "D": 18, "E": 18, "F": 70})

ws["B2"] = "Enterprise Roles & Daily Rates"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: This is the Enterprise team. SMB roles + 4 enterprise specialists. The 'Internal cost' is what each person costs us per day."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

hdr_r = 5
hdrs = ["Role", "Seniority", "Internal cost / day (USD)", "Bill rate / day (USD)", "Notes"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# (role, seniority, internal_cost, bill_rate, note)
roles = [
    # SMB-shared roles (same daily rates)
    ("Engagement / Programme Manager", "Senior", 1000, 1750, "Single point of contact to e&; weekly governance; commercial owner."),
    ("Solution Architect (Lead)", "Principal", 1100, 2100, "Owns reference architecture, multi-tenant design, latency budget, capability matrix."),
    ("AI / LLM Engineer — Orchestration & Reasoning", "Senior", 900, 1700, "Agent graph, FSM, tool-calling, memory, guardrails, Portkey integration."),
    ("Voice / Conversational Engineer", "Senior", 850, 1650, "Telephony, STT/TTS, WhatsApp/Web channels, Arabic/English NLP."),
    ("Data & Integration Engineer", "Senior", 900, 1550, "Connectors (CRM, calendar, payments), e& billing API, vector store, knowledge ingestion."),
    ("Frontend / UX Engineer", "Mid", 750, 1300, "Next.js tenant back office, onboarding, admin console, agent configuration UI."),
    ("DevOps / Platform Engineer", "Senior", 950, 1550, "Helm/Terraform, CI/CD, multi-tenant K8s, observability stack, secrets, cost guardrails."),
    ("QA Engineer", "Mid", 700, 1200, "Test packs, synthetic data, automated regression, voice/chat acceptance tests."),
    ("Security / Compliance Lead", "Senior", 1050, 1750, "Tenant isolation, NESA / TDRA / ADDA controls, pen-test coordination."),
    ("Customer Success / Beta Ops Lead", "Mid", 600, 950, "Pilot onboarding, feedback loops, runbooks, support escalation."),
    # Enterprise-specific roles (added)
    ("ERP / Systems Integration Architect", "Principal", 1200, 2200, "Oracle Fusion / SAP S/4HANA / Epic / Temenos / Dynamics deep integration. Critical for Tier 1/2/3 enterprise builds."),
    ("Regulatory & Compliance SME", "Senior", 1100, 1800, "DHA / CBUAE / NESA / ADDA / FTA. Maps regulation to controls; named on every Tier 2/3 audit trail."),
    ("Industry SME (per vertical)", "Senior", 1000, 1700, "Healthcare / Banking / Energy / Government domain expert. Spec authoring + acceptance testing."),
    ("ML / Fine-Tuning Engineer", "Senior", 950, 1700, "Domain fine-tuning (medical Arabic, legal Arabic, financial Arabic) on sovereign GPU. RAG quality."),
    ("Hardware / On-Prem Infra Engineer (Tier 3)", "Senior", 1000, 1700, "On-customer-site GPU cluster install, networking, storage, NVAIE; coordinates with Help AG for security wrap."),
]
for i, (role, sen, ic, br, note) in enumerate(roles):
    r = hdr_r + 1 + i
    ws.cell(row=r, column=2, value=role); ws.cell(row=r, column=2).font = F_LBL; ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=3, value=sen); ws.cell(row=r, column=3).font = F_TXT; ws.cell(row=r, column=3).alignment = ALIGN_C; ws.cell(row=r, column=3).border = BORDER
    icc = ws.cell(row=r, column=4, value=ic); style_input(icc); icc.number_format = "$#,##0"
    brc = ws.cell(row=r, column=5, value=br); style_input(brc); brc.number_format = "$#,##0"
    nc = ws.cell(row=r, column=6, value=note); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

total_r = hdr_r + 1 + len(roles) + 1
ws.cell(row=total_r, column=2, value=f"Total roles in plan: {len(roles)} (10 SMB-shared + 5 enterprise-specific)").font = F_LBL


# =============================================================================
# SHEET 3 — Wave E0 — Enterprise Platform Extensions
# =============================================================================
ws = wb.create_sheet("3. Wave E0 - Platform")
set_widths(ws, {"A": 3, "B": 46, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 16, "I": 50})

ws["B2"] = "Wave E0 — Enterprise Platform Extensions  (P1 + P2 layers on top of SMB platform)"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: One-off build of the enterprise-readiness layer that Tier 1/2/3 deals reuse. RBAC, audit trail, trust tiers, ERP framework, fine-tuning sub-platform. ~12 weeks across 5 phases."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

phases = [
    "EP1 — Mobilisation & SDD",
    "EP2 — Trust Tiers + RBAC + Audit",
    "EP3 — ERP/CRM/HIS Integration Framework",
    "EP4 — Fine-Tuning Sub-Platform",
    "EP5 — Hardening, Pen-Test, GA",
]
phase_weeks = [2, 3, 3, 2, 2]  # 12 weeks total

# p-weeks per role per phase (heuristic, lighter than SMB Wave 1 because we're extending, not building from zero)
role_names = [r[0] for r in roles]
n_roles = len(role_names)

# Sizing: 15 roles × 5 phases. Many SMB-shared roles have lower allocation here.
sizing_e0 = {
    0:  [1, 1, 1, 1, 1],         # PM (1 person each phase)
    1:  [2, 2, 2, 1, 1],         # Sol Arch (very heavy on platform extensions)
    2:  [1, 2, 2, 2, 1],         # AI/LLM Eng
    3:  [0, 0, 0, 0, 0],         # Voice (not in platform extension)
    4:  [1, 2, 3, 1, 1],         # Data/Integration
    5:  [1, 2, 2, 1, 1],         # Frontend (admin console for trust tiers, audit viewer, RBAC UI)
    6:  [2, 3, 2, 2, 2],         # DevOps (heavy — multi-tenant, namespace isolation, K8s policies)
    7:  [1, 2, 2, 1, 2],         # QA
    8:  [2, 3, 2, 2, 2],         # Security/Compliance Lead (very heavy — trust tiers + audit)
    9:  [0, 0, 0, 0, 0],         # CS (not in platform build)
    10: [1, 1, 3, 1, 1],         # ERP Integration Architect (peak on EP3)
    11: [1, 2, 2, 1, 2],         # Regulatory & Compliance SME
    12: [0, 0, 0, 0, 0],         # Industry SME (per vertical — billed to deals not platform)
    13: [0, 0, 0, 2, 1],         # Fine-Tuning Eng (peak on EP4)
    14: [0, 0, 0, 0, 0],         # Hardware Infra (Tier 3 only)
}

# Header — phase labels
hdr_r = 5
ws.cell(row=hdr_r, column=2, value="Role").font = F_HDR
ws.cell(row=hdr_r, column=2).fill = FILL_HDR; ws.cell(row=hdr_r, column=2).alignment = ALIGN_C; ws.cell(row=hdr_r, column=2).border = BORDER
for i, p in enumerate(phases):
    c = ws.cell(row=hdr_r, column=3 + i, value=p)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER
ws.cell(row=hdr_r, column=8, value="Total p-weeks").font = F_HDR
ws.cell(row=hdr_r, column=8).fill = FILL_HDR; ws.cell(row=hdr_r, column=8).alignment = ALIGN_C; ws.cell(row=hdr_r, column=8).border = BORDER
ws.cell(row=hdr_r, column=9, value="Notes").font = F_HDR
ws.cell(row=hdr_r, column=9).fill = FILL_HDR; ws.cell(row=hdr_r, column=9).alignment = ALIGN_C; ws.cell(row=hdr_r, column=9).border = BORDER

# Phase duration row
ws.cell(row=6, column=2, value="Phase duration (weeks)").font = F_LBL
ws.cell(row=6, column=2).alignment = ALIGN_L; ws.cell(row=6, column=2).border = BORDER
for i, w in enumerate(phase_weeks):
    cell = ws.cell(row=6, column=3 + i, value=w)
    style_input(cell); cell.number_format = "0"
ws.cell(row=6, column=8, value=f"=SUM(C6:G6)"); style_total(ws.cell(row=6, column=8)); ws.cell(row=6, column=8).number_format = "0"
ws.cell(row=6, column=9).border = BORDER

note_e0 = {
    0: "Engagement coordination across the 12-week build.",
    1: "Architects own trust tier framework, namespace design, audit pipeline, ERP integration patterns.",
    2: "Tools-calling adapters for ERP/CRM/HIS, SSO providers, audit emitters.",
    3: "—",
    4: "ERP/CRM/HIS connector framework (Oracle, SAP, Epic, Temenos, Dynamics).",
    5: "Admin console: trust tier UI, RBAC, audit viewer, fine-tune dashboard.",
    6: "Tenant isolation, namespace-per-tenant (T2), kernel-sandboxed runtime (T3 prep).",
    7: "Security testing, compliance test pack, regression for trust-tier transitions.",
    8: "Owns NESA / ISO 27017/27018 / SOC 2 Type II uplift; engages pen-test partner.",
    9: "—",
    10: "Owns ERP integration patterns library — reusable across every Tier 1/2/3 deal.",
    11: "Maps DHA / CBUAE / NESA / ADDA controls to platform features; signs off on audit evidence.",
    12: "—",
    13: "Stands up fine-tuning sub-platform — sovereign training jobs on G42 GPU.",
    14: "—",
}

# Role rows: p-weeks × cost
for i, role in enumerate(role_names):
    r = hdr_r + 2 + i  # rows 7..21
    cl = ws.cell(row=r, column=2, value=role); cl.font = F_LBL; cl.alignment = ALIGN_L; cl.border = BORDER
    pweeks = sizing_e0[i]
    for j, pw in enumerate(pweeks):
        cell = ws.cell(row=r, column=3 + j, value=pw)
        style_input(cell); cell.number_format = "0"
    tot = ws.cell(row=r, column=8, value=f"=SUM(C{r}:G{r})")
    style_total(tot); tot.number_format = "0"
    nc = ws.cell(row=r, column=9, value=note_e0[i]); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# Total p-weeks row
total_pw_row = hdr_r + 2 + n_roles  # row 22
ws.cell(row=total_pw_row, column=2, value="Total person-weeks per phase").font = F_TOT
ws.cell(row=total_pw_row, column=2).fill = FILL_TOTAL; ws.cell(row=total_pw_row, column=2).border = BORDER
for j in range(5):
    col_letter = get_column_letter(3 + j)
    cell = ws.cell(row=total_pw_row, column=3 + j, value=f"=SUM({col_letter}{hdr_r+2}:{col_letter}{total_pw_row-1})")
    style_total(cell); cell.number_format = "0"
gt = ws.cell(row=total_pw_row, column=8, value=f"=SUM(H{hdr_r+2}:H{total_pw_row-1})")
style_total(gt); gt.number_format = "0"; gt.font = Font(name="Calibri", size=11, bold=True, color=NAVY)

# Cost section
cost_lbl_row = total_pw_row + 2
ws.cell(row=cost_lbl_row, column=2, value="Personnel cost (USD)").font = F_H1
WD_REF = "'1. Cover & Assumptions'!$C$24"  # working days/week

cost_row = cost_lbl_row + 1
ws.cell(row=cost_row, column=2, value="Personnel cost — sum across phases").font = F_LBL
ws.cell(row=cost_row, column=2).alignment = ALIGN_L; ws.cell(row=cost_row, column=2).border = BORDER
parts = []
for i in range(n_roles):
    role_row = hdr_r + 2 + i
    sheet2_row = 6 + i  # Sheet 2 internal cost lookup
    parts.append(f"H{role_row}*'2. Roles & Rates'!D{sheet2_row}")
formula = f"=({'+'.join(parts)})*{WD_REF}"
cost_cell = ws.cell(row=cost_row, column=8, value=formula)
style_total(cost_cell); cost_cell.number_format = "$#,##0"
cost_cell.font = Font(name="Calibri", size=11, bold=True, color=NAVY)

# Other costs — platform build other costs
other_row = cost_row + 1
ws.cell(row=other_row, column=2, value="Other costs (cloud, LLM, security tools, pen-test)").font = F_LBL
ws.cell(row=other_row, column=2).alignment = ALIGN_L; ws.cell(row=other_row, column=2).border = BORDER
oc = ws.cell(row=other_row, column=8, value=85000)  # ~$85K for 12-week extension build (lighter than Wave 1 because shorter & no GA marketing)
style_input(oc); oc.number_format = "$#,##0"
ws.cell(row=other_row, column=9, value="Cloud dev/staging, LLM gateway, observability, pen-test ($25K), security tools, contingency cushion. Editable.").font = F_NOTE
ws.cell(row=other_row, column=9).border = BORDER

# Direct cost
direct_row = other_row + 1
ws.cell(row=direct_row, column=2, value="Direct cost — sum").font = F_TOT
ws.cell(row=direct_row, column=2).fill = FILL_TOTAL; ws.cell(row=direct_row, column=2).border = BORDER
dc = ws.cell(row=direct_row, column=8, value=f"=H{cost_row}+H{other_row}")
style_total(dc); dc.number_format = "$#,##0"

# Contingency
cont_row = direct_row + 1
ws.cell(row=cont_row, column=2, value="Contingency (12% — moderate complexity)").font = F_LBL
ws.cell(row=cont_row, column=2).alignment = ALIGN_L; ws.cell(row=cont_row, column=2).border = BORDER
cc = ws.cell(row=cont_row, column=8, value=f"=H{direct_row}*0.12")
style_formula(cc); cc.number_format = "$#,##0"

# Total fully loaded
loaded_row = cont_row + 1
ws.cell(row=loaded_row, column=2, value="Total fully-loaded cost").font = F_TOT
ws.cell(row=loaded_row, column=2).fill = FILL_TOTAL; ws.cell(row=loaded_row, column=2).border = BORDER
fl = ws.cell(row=loaded_row, column=8, value=f"=H{direct_row}+H{cont_row}")
style_total(fl); fl.number_format = "$#,##0"
fl.font = Font(name="Calibri", size=11, bold=True, color=NAVY)


# =============================================================================
# SHEET 4 — Tier 1: Adapted SMB Agents (delta only)
# =============================================================================
ws = wb.create_sheet("4. Tier 1 - Adapted Agents")
set_widths(ws, {"A": 3, "B": 36, "C": 16, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14, "K": 50})

ws["B2"] = "Tier 1 — Adapted SMB Agents (Enterprise Edition)"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: Each enterprise customer gets adapted versions of the SMB agents — same agent IP, but with multi-branch routing, deep ERP/CRM/HIS connectors, compliance logging, fine-tuning. We DON'T rebuild the agent — we configure & extend it. That is why the cost is only ~$30-60K per agent."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

# Six adapted agents (one per SMB agent)
adapted = [
    "Customer Agent (Enterprise)",
    "Sales Agent (Enterprise)",
    "Comms Hub (Enterprise)",
    "Finance Agent (Enterprise)",
    "Ops Agent (Enterprise)",
    "People Agent (Enterprise)",
]

# p-weeks per role for ONE adapted agent (delta work only — 6-9 weeks duration per agent)
# Rationale: SMB agent already exists. Enterprise delta = config + ERP connector + compliance + fine-tune + UAT.
adapted_pweeks = {
    0:  [1, 1, 1, 1, 1, 1],   # PM
    1:  [1, 1, 1, 1, 1, 1],   # Sol Arch (light — review & sign-off)
    2:  [2, 2, 2, 2, 2, 2],   # AI/LLM Eng (prompt tuning, tools, evals)
    3:  [1, 0, 2, 0, 0, 0],   # Voice (Customer + Comms only)
    4:  [3, 3, 3, 4, 3, 4],   # Data/Integration (heaviest — ERP/CRM/HIS connector)
    5:  [1, 1, 1, 1, 1, 1],   # Frontend (light tweaks)
    6:  [1, 1, 1, 1, 1, 1],   # DevOps
    7:  [1, 1, 1, 2, 1, 2],   # QA (Finance/People = stronger compliance test)
    8:  [1, 1, 1, 2, 1, 2],   # Security (Finance/People = PII heavier)
    9:  [1, 1, 1, 1, 1, 1],   # CS
    10: [2, 2, 2, 3, 2, 3],   # ERP Architect — KEY enterprise role (Finance/People = SAP/Oracle HR heaviest)
    11: [1, 1, 1, 2, 1, 2],   # Compliance SME (Finance/People heaviest)
    12: [2, 2, 2, 2, 2, 2],   # Industry SME (per vertical)
    13: [1, 1, 1, 1, 1, 1],   # Fine-Tuning Eng (light per-agent fine-tune)
    14: [0, 0, 0, 0, 0, 0],   # Hardware infra (only Tier 3)
}

note_adapted = {
    0: "Engagement & e&-customer coordination.",
    1: "Solution architect signs off integration design + trust tier mapping.",
    2: "Enterprise prompt tuning + new tools per ERP/CRM/HIS context.",
    3: "Voice only on Customer Agent + Comms Hub (multi-branch IVR).",
    4: "Heaviest role — connector to customer's ERP/CRM/HIS instance.",
    5: "Light UI tweaks — multi-branch routing, custom branding.",
    6: "Tenant onboarding into customer's environment + observability hooks.",
    7: "Per-agent acceptance tests + compliance test pack.",
    8: "Compliance logging + audit trail wiring + RBAC config.",
    9: "Pilot user onboarding + hypercare for 30 days.",
    10: "Critical role — Oracle Fusion / SAP S/4HANA / Epic / Temenos / Dynamics integration.",
    11: "Maps customer-specific compliance (DHA / CBUAE / NESA / ADDA) to controls.",
    12: "Healthcare / Banking / Energy / Real Estate / Government domain expert.",
    13: "Light per-agent fine-tune (industry vocab, customer-specific patterns).",
    14: "—",
}

# Header
hdr_r = 5
hdrs = ["Role", "Cost / day"] + adapted + ["Total p-weeks", "Notes"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

for i, role in enumerate(role_names):
    r = hdr_r + 1 + i  # 6..20
    cl = ws.cell(row=r, column=2, value=role); cl.font = F_LBL; cl.alignment = ALIGN_L; cl.border = BORDER
    sheet2_row = 6 + i
    cd = ws.cell(row=r, column=3, value=f"='2. Roles & Rates'!D{sheet2_row}")
    style_formula(cd); cd.number_format = "$#,##0"
    pweeks = adapted_pweeks[i]
    for j, pw in enumerate(pweeks):
        cell = ws.cell(row=r, column=4 + j, value=pw)
        style_input(cell); cell.number_format = "0"
    tot = ws.cell(row=r, column=10, value=f"=SUM(D{r}:I{r})")
    style_total(tot); tot.number_format = "0"
    nc = ws.cell(row=r, column=11, value=note_adapted[i]); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# Total p-weeks per agent row
tot_pw_row = hdr_r + 1 + n_roles
ws.cell(row=tot_pw_row, column=2, value="Total person-weeks per agent").font = F_TOT
ws.cell(row=tot_pw_row, column=2).fill = FILL_TOTAL; ws.cell(row=tot_pw_row, column=2).border = BORDER
for j in range(6):
    col_letter = get_column_letter(4 + j)
    cell = ws.cell(row=tot_pw_row, column=4 + j, value=f"=SUM({col_letter}{hdr_r+1}:{col_letter}{tot_pw_row-1})")
    style_total(cell); cell.number_format = "0"
gt = ws.cell(row=tot_pw_row, column=10, value=f"=SUM(J{hdr_r+1}:J{tot_pw_row-1})")
style_total(gt); gt.number_format = "0"

# Cost block
cost_lbl_row = tot_pw_row + 2
ws.cell(row=cost_lbl_row, column=2, value="Cost per agent (USD)").font = F_H1
hdr_r2 = cost_lbl_row + 1
hdrs2 = ["Item"] + adapted + ["Avg per agent"]
for i, h in enumerate(hdrs2):
    c = ws.cell(row=hdr_r2, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Personnel cost per agent
pers_row = hdr_r2 + 1
ws.cell(row=pers_row, column=2, value="Personnel cost").font = F_LBL
ws.cell(row=pers_row, column=2).alignment = ALIGN_L; ws.cell(row=pers_row, column=2).border = BORDER
for j in range(6):
    col_pw = get_column_letter(4 + j)
    parts = []
    for i in range(n_roles):
        role_row = hdr_r + 1 + i
        parts.append(f"{col_pw}{role_row}*$C{role_row}")
    formula = f"=({'+'.join(parts)})*{WD_REF}"
    cell = ws.cell(row=pers_row, column=4 + j, value=formula)
    style_total(cell); cell.number_format = "$#,##0"
avg_cell = ws.cell(row=pers_row, column=10, value=f"=AVERAGE(D{pers_row}:I{pers_row})")
style_total(avg_cell); avg_cell.number_format = "$#,##0"
avg_cell.font = Font(name="Calibri", size=11, bold=True, color=NAVY)

# Other costs per agent (light — ~$8K cloud sandbox + integration test sandbox + fine-tune compute)
oc_row = pers_row + 1
ws.cell(row=oc_row, column=2, value="Other costs (per agent — sandbox, fine-tune compute, integration env)").font = F_LBL
ws.cell(row=oc_row, column=2).alignment = ALIGN_L; ws.cell(row=oc_row, column=2).border = BORDER
for j in range(6):
    cell = ws.cell(row=oc_row, column=4 + j, value=8000)
    style_input(cell); cell.number_format = "$#,##0"
avg_oc = ws.cell(row=oc_row, column=10, value=f"=AVERAGE(D{oc_row}:I{oc_row})")
style_total(avg_oc); avg_oc.number_format = "$#,##0"

# Total per agent
tot_row = oc_row + 1
ws.cell(row=tot_row, column=2, value="Total cost per agent (Tier 1)").font = F_TOT
ws.cell(row=tot_row, column=2).fill = FILL_TOTAL; ws.cell(row=tot_row, column=2).border = BORDER
for j in range(6):
    col_letter = get_column_letter(4 + j)
    cell = ws.cell(row=tot_row, column=4 + j, value=f"={col_letter}{pers_row}+{col_letter}{oc_row}")
    style_total(cell); cell.number_format = "$#,##0"
avg_tot = ws.cell(row=tot_row, column=10, value=f"=AVERAGE(D{tot_row}:I{tot_row})")
style_total(avg_tot); avg_tot.number_format = "$#,##0"
avg_tot.font = Font(name="Calibri", size=11, bold=True, color=NAVY)

ws.cell(row=tot_row + 2, column=2, value="↑ Note: cost EXCLUDES Wave E0 platform extensions (one-off in Sheet 3) and any Help AG security wrap (Tier 3 only).").font = F_NOTE


# =============================================================================
# SHEET 5 — Tier 2: Custom AI Solutions
# =============================================================================
ws = wb.create_sheet("5. Tier 2 - Custom AI")
set_widths(ws, {"A": 3, "B": 36, "C": 16, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 50})

ws["B2"] = "Tier 2 — Custom AI Solutions  (incl. AI Contact Centre Software for Enterprises)"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: Bigger custom AI solutions that go beyond the marketplace agents. AI Contact Centre is the headline solution — full CCaaS with voice, WhatsApp, web, agent assist, supervisor analytics, deep ERP/CRM/HIS integration. ~16-24 weeks per project."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

# Five custom solutions
custom = [
    ("AI Contact Centre Software (Enterprise)", "Full CCaaS — voice + WhatsApp + web + agent assist + supervisor + analytics."),
    ("Document Intelligence Suite", "Inbound docs (forms, contracts, invoices, medical records) classified, extracted, routed."),
    ("Approval Orchestration Platform", "Multi-step approvals with policy + RAG + human-in-loop (procurement, expenses, credit)."),
    ("Predictive Maintenance Agent", "IoT/SCADA telemetry → anomaly detection + work-order automation + technician dispatch."),
    ("Security & Compliance Agent", "AI SOC analyst, log triage, compliance evidence, regulator reporting."),
]

# Per-project p-weeks per role (CCaaS is heaviest, others are lighter)
# Customer Care Centre estimated ~22 weeks; Doc Intel ~18 weeks; Approvals ~16 weeks; Pred Maint ~20 weeks; Security ~18 weeks.
custom_pweeks = {
    0:  [4, 4, 3, 4, 4],            # PM
    1:  [3, 3, 3, 3, 3],            # Sol Arch
    2:  [10, 8, 8, 8, 8],           # AI/LLM Eng
    3:  [12, 0, 0, 0, 0],           # Voice (CCaaS only — STT/TTS/IVR/agent assist)
    4:  [10, 8, 6, 10, 6],          # Data/Integration
    5:  [6, 6, 6, 4, 4],            # Frontend
    6:  [4, 3, 3, 4, 3],            # DevOps
    7:  [5, 5, 4, 5, 4],            # QA
    8:  [3, 3, 4, 3, 6],            # Security (peak on Security & Compliance Agent)
    9:  [4, 3, 3, 3, 3],            # CS / Beta Ops
    10: [4, 4, 4, 4, 3],            # ERP Architect
    11: [3, 3, 4, 2, 5],            # Compliance SME (peak Security)
    12: [3, 3, 3, 4, 2],            # Industry SME
    13: [3, 4, 2, 3, 2],            # Fine-Tuning Eng (Doc Intel needs more)
    14: [0, 0, 0, 0, 0],            # Hardware (Tier 3 only)
}

note_custom = {
    0: "Engagement coordination across the 16-22 week build.",
    1: "Architect owns SDD, integration design, latency budget, fallback strategy.",
    2: "Heaviest role — orchestration, prompt engineering, evals, tool calling.",
    3: "CCaaS only — STT/TTS, real-time transcription, agent assist, post-call summary.",
    4: "ERP/CRM/HIS integration; CCaaS = CRM screen-pop + ticket creation.",
    5: "Supervisor/agent UI; CCaaS = wallboards, queue mgmt; Doc Intel = review UI.",
    6: "Per-project deployment + observability + cost guardrails.",
    7: "Heavy test coverage — voice acceptance, regression, regulator scenarios.",
    8: "Security & Compliance Agent peaks here; others moderate.",
    9: "Pilot user onboarding + 30-day hypercare per project.",
    10: "Deep integration to customer ERP — CCaaS needs ticketing system bridge.",
    11: "Compliance Agent needs heaviest SME involvement — full regulatory mapping.",
    12: "Healthcare / Banking / Energy / Government / Manufacturing domain expert.",
    13: "Doc Intel: invoice/medical record fine-tune. CCaaS: industry-specific intent.",
    14: "—",
}

hdr_r = 5
hdrs = ["Role", "Cost / day"] + [c[0] for c in custom] + ["Notes"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

for i, role in enumerate(role_names):
    r = hdr_r + 1 + i
    cl = ws.cell(row=r, column=2, value=role); cl.font = F_LBL; cl.alignment = ALIGN_L; cl.border = BORDER
    sheet2_row = 6 + i
    cd = ws.cell(row=r, column=3, value=f"='2. Roles & Rates'!D{sheet2_row}")
    style_formula(cd); cd.number_format = "$#,##0"
    pweeks = custom_pweeks[i]
    for j, pw in enumerate(pweeks):
        cell = ws.cell(row=r, column=4 + j, value=pw)
        style_input(cell); cell.number_format = "0"
    nc = ws.cell(row=r, column=9, value=note_custom[i]); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# Total p-weeks per project
tot_pw_row = hdr_r + 1 + n_roles
ws.cell(row=tot_pw_row, column=2, value="Total person-weeks per project").font = F_TOT
ws.cell(row=tot_pw_row, column=2).fill = FILL_TOTAL; ws.cell(row=tot_pw_row, column=2).border = BORDER
for j in range(5):
    col_letter = get_column_letter(4 + j)
    cell = ws.cell(row=tot_pw_row, column=4 + j, value=f"=SUM({col_letter}{hdr_r+1}:{col_letter}{tot_pw_row-1})")
    style_total(cell); cell.number_format = "0"

# Cost section
cost_lbl_row = tot_pw_row + 2
ws.cell(row=cost_lbl_row, column=2, value="Cost per project (USD)").font = F_H1
hdr_r2 = cost_lbl_row + 1
hdrs2 = ["Item"] + [c[0] for c in custom] + ["Avg per project"]
for i, h in enumerate(hdrs2):
    c = ws.cell(row=hdr_r2, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Personnel cost per project
pers_row = hdr_r2 + 1
ws.cell(row=pers_row, column=2, value="Personnel cost").font = F_LBL
ws.cell(row=pers_row, column=2).alignment = ALIGN_L; ws.cell(row=pers_row, column=2).border = BORDER
for j in range(5):
    col_pw = get_column_letter(4 + j)
    parts = []
    for i in range(n_roles):
        role_row = hdr_r + 1 + i
        parts.append(f"{col_pw}{role_row}*$C{role_row}")
    formula = f"=({'+'.join(parts)})*{WD_REF}"
    cell = ws.cell(row=pers_row, column=4 + j, value=formula)
    style_total(cell); cell.number_format = "$#,##0"
avg_pers = ws.cell(row=pers_row, column=9, value=f"=AVERAGE(D{pers_row}:H{pers_row})")
style_total(avg_pers); avg_pers.number_format = "$#,##0"

# Other costs (cloud, voice APIs, LLM, sandbox, integration env)
oc_row = pers_row + 1
ws.cell(row=oc_row, column=2, value="Other costs (cloud, LLM, voice APIs, sandbox)").font = F_LBL
ws.cell(row=oc_row, column=2).alignment = ALIGN_L; ws.cell(row=oc_row, column=2).border = BORDER
# CCaaS = $40K (voice APIs heaviest), others = $20K
oc_per_project = [40000, 25000, 20000, 25000, 20000]
for j, oc in enumerate(oc_per_project):
    cell = ws.cell(row=oc_row, column=4 + j, value=oc)
    style_input(cell); cell.number_format = "$#,##0"
avg_oc = ws.cell(row=oc_row, column=9, value=f"=AVERAGE(D{oc_row}:H{oc_row})")
style_total(avg_oc); avg_oc.number_format = "$#,##0"

# Total per project
tot_row = oc_row + 1
ws.cell(row=tot_row, column=2, value="Total cost per project (Tier 2)").font = F_TOT
ws.cell(row=tot_row, column=2).fill = FILL_TOTAL; ws.cell(row=tot_row, column=2).border = BORDER
for j in range(5):
    col_letter = get_column_letter(4 + j)
    cell = ws.cell(row=tot_row, column=4 + j, value=f"={col_letter}{pers_row}+{col_letter}{oc_row}")
    style_total(cell); cell.number_format = "$#,##0"
avg_tot = ws.cell(row=tot_row, column=9, value=f"=AVERAGE(D{tot_row}:H{tot_row})")
style_total(avg_tot); avg_tot.number_format = "$#,##0"
avg_tot.font = Font(name="Calibri", size=11, bold=True, color=NAVY)

# Solution descriptions
desc_row = tot_row + 3
ws.cell(row=desc_row, column=2, value="Solution descriptions").font = F_H1
for i, (name, desc) in enumerate(custom):
    r = desc_row + 1 + i
    ws.cell(row=r, column=2, value=name).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    ws.cell(row=r, column=3, value=desc).font = F_NOTE
    ws.cell(row=r, column=3).alignment = ALIGN_L; ws.cell(row=r, column=3).border = BORDER


# =============================================================================
# SHEET 6 — Tier 3: Sovereign / On-Prem
# =============================================================================
ws = wb.create_sheet("6. Tier 3 - Sovereign OnPrem")
set_widths(ws, {"A": 3, "B": 38, "C": 16, "D": 16, "E": 16, "F": 16, "G": 50})

ws["B2"] = "Tier 3 — Sovereign & On-Prem Deployments"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: Customer-site deployments (gov, banks, healthcare regulators). Higher complexity — customer-specific architecture, networking, security, audit. Pre-prepared HPC RA BoMs (Sheet 10) let us spin up vendor-agnostic BoMs in days. Help AG security wrap included."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

# Cost components for an average Tier 3 project
ws["B5"] = "Cost stack — typical Tier 3 sovereign / on-prem deployment"
ws["B5"].font = F_H1

hdr_r = 6
for i, h in enumerate(["Cost component", "Value (USD)", "Notes"]):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Total Tier 2 personnel cost (avg project) — Sheet 5 row tot_row col I (col 9)
# Tier 3 = Tier 2 baseline build + uplift for on-prem complexity
# We compute this by referencing Tier 2 avg + a multiplier.
# Tier 2 personnel avg cell — easier to compute fresh. Tier 2 sheet rows: pers_row, oc_row, tot_row (these are local to that sheet).
# For Tier 3, we model from scratch:
tier3_rows = [
    ("Base custom AI solution build (avg Tier 2 personnel)", "='5. Tier 2 - Custom AI'!I27", "References Tier 2 average personnel cost (Doc Intel / Approvals / Pred Maint median build)."),
    ("On-prem uplift — installation, networking, identity (8 wks Hardware Eng + Sec Lead)", "=8*'1. Cover & Assumptions'!$C$24*('2. Roles & Rates'!D20+'2. Roles & Rates'!D14)", "8 weeks × 5 days × (Hardware Eng + Security Lead). Customer-site networking, AD/SSO integration, hardening."),
    ("Customer-specific architecture & SDD addendum (3 wks Sol Arch + ERP Arch)", "=3*'1. Cover & Assumptions'!$C$24*('2. Roles & Rates'!D7+'2. Roles & Rates'!D16)", "3-week architecture deep-dive with customer IT + procurement before build starts."),
    ("On-prem environment & integration sandbox", 25000, "Bare-metal lab in AIdeology / e& staging that mirrors customer footprint for testing."),
    ("Help AG managed security wrap", 75000, "Third-party security partner — pen-test, SOC integration, NESA evidence. Indicative — verify per project."),
    ("HPC BoM coordination & vendor liaison (allowance)", 25000, "AIdeology engineering time to liaise with chosen vendor (Cisco/Dell/HPE/Lenovo/Supermicro) on BoM, lead times, install."),
    ("Travel & on-site presence (4 trips × 5 people × 5 days)", 60000, "On-site mobilisation, install, UAT, go-live."),
    ("Other costs (cloud, LLM, sandbox uplift)", 30000, "Higher than Tier 2 — air-gapped test env, dual-stack."),
]
for i, (label, val, note) in enumerate(tier3_rows):
    r = hdr_r + 1 + i
    ws.cell(row=r, column=2, value=label).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    cell = ws.cell(row=r, column=3, value=val)
    if isinstance(val, (int, float)):
        style_input(cell)
    else:
        style_formula(cell)
    cell.number_format = "$#,##0"
    nc = ws.cell(row=r, column=4, value=note); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# Subtotal direct cost
sub_row = hdr_r + 1 + len(tier3_rows)
ws.cell(row=sub_row, column=2, value="Direct cost — sum").font = F_TOT
ws.cell(row=sub_row, column=2).fill = FILL_TOTAL; ws.cell(row=sub_row, column=2).border = BORDER
sc = ws.cell(row=sub_row, column=3, value=f"=SUM(C{hdr_r+1}:C{sub_row-1})")
style_total(sc); sc.number_format = "$#,##0"
sc.font = Font(name="Calibri", size=11, bold=True, color=NAVY)

# Contingency
cont_row = sub_row + 1
ws.cell(row=cont_row, column=2, value="Contingency (15% — high-complexity buffer)").font = F_LBL
ws.cell(row=cont_row, column=2).alignment = ALIGN_L; ws.cell(row=cont_row, column=2).border = BORDER
cc = ws.cell(row=cont_row, column=3, value=f"=C{sub_row}*'1. Cover & Assumptions'!$C$37")
style_formula(cc); cc.number_format = "$#,##0"

# Total fully loaded
loaded_row = cont_row + 1
ws.cell(row=loaded_row, column=2, value="Total fully-loaded cost — typical Tier 3 project").font = F_TOT
ws.cell(row=loaded_row, column=2).fill = FILL_TOTAL; ws.cell(row=loaded_row, column=2).border = BORDER
fl = ws.cell(row=loaded_row, column=3, value=f"=C{sub_row}+C{cont_row}")
style_total(fl); fl.number_format = "$#,##0"
fl.font = Font(name="Calibri", size=12, bold=True, color=NAVY)

# Hardware BoM note
note_row = loaded_row + 3
ws.cell(row=note_row, column=2, value="↑ Important — Hardware BoM is NOT in our cost above").font = F_H1
hw_notes = [
    "The HPC compute infrastructure (GPU cluster, networking, storage) is sold as a pass-through line via the chosen OEM (Cisco / Dell / HPE / Lenovo / Supermicro).",
    "AIdeology coordinates the BoM (allowance included above), sizes the cluster against customer workload, but does not carry hardware revenue. e& or the OEM owns that line.",
    "See Sheet 10 — HPC RA Library — for ten reference architectures from five vendors with sizing guidance (4-128 nodes, RTX PRO / HGX / B300 / NVL72).",
    "Optional: AIdeology can bundle the BoM into the price (with a 5-10% margin) — discuss per deal. Default model: hardware sold separately by OEM.",
]
for i, line in enumerate(hw_notes):
    ws.cell(row=note_row + 1 + i, column=2, value=line).font = F_NOTE


# =============================================================================
# SHEET 7 — Custom Bespoke
# =============================================================================
ws = wb.create_sheet("7. Custom Bespoke")
set_widths(ws, {"A": 3, "B": 44, "C": 14, "D": 14, "E": 14, "F": 14, "G": 50})

ws["B2"] = "Custom Bespoke Solutions  (AIdeology unique builds)"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: When an enterprise needs something that doesn't fit Tier 1/2/3 — a unique vertical AI, a moonshot R&D project, or new platform IP. Priced T&M or fixed-bid, premium margin (50%) because of unique IP and execution risk."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

# Sample bespoke project sizing (e.g., bespoke regulator-grade AI, multi-LLM router, vertical agent factory)
ws["B5"] = "Reference bespoke project sizing (~22 weeks, 8-person team)"
ws["B5"].font = F_H1

hdr_r = 6
hdrs = ["Role", "Cost / day", "p-weeks", "Personnel cost", "Notes"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Bespoke team — heavier on architects + AI eng + ML/Fine-tune; lighter on voice/CS
bespoke_pweeks = {
    0: (5,  "PM"),
    1: (6,  "Sol Arch"),
    2: (15, "AI/LLM Eng (heavy R&D)"),
    3: (3,  "Voice (if needed)"),
    4: (10, "Data/Integration"),
    5: (8,  "Frontend"),
    6: (4,  "DevOps"),
    7: (6,  "QA"),
    8: (4,  "Security/Compliance"),
    9: (3,  "CS / Beta Ops"),
    10:(3,  "ERP Architect (if needed)"),
    11:(3,  "Compliance SME"),
    12:(4,  "Industry SME"),
    13:(8,  "Fine-Tuning Eng (heavy)"),
    14:(0,  "Hardware (only if on-prem)"),
}

for i, role in enumerate(role_names):
    r = hdr_r + 1 + i
    cl = ws.cell(row=r, column=2, value=role); cl.font = F_LBL; cl.alignment = ALIGN_L; cl.border = BORDER
    sheet2_row = 6 + i
    cd = ws.cell(row=r, column=3, value=f"='2. Roles & Rates'!D{sheet2_row}")
    style_formula(cd); cd.number_format = "$#,##0"
    pw, _ = bespoke_pweeks[i]
    pwc = ws.cell(row=r, column=4, value=pw); style_input(pwc); pwc.number_format = "0"
    pers = ws.cell(row=r, column=5, value=f"=C{r}*D{r}*'1. Cover & Assumptions'!$C$24")
    style_formula(pers); pers.number_format = "$#,##0"
    nc = ws.cell(row=r, column=7, value=bespoke_pweeks[i][1]); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# Total
tot_row = hdr_r + 1 + n_roles
ws.cell(row=tot_row, column=2, value="Total — bespoke build personnel").font = F_TOT
ws.cell(row=tot_row, column=2).fill = FILL_TOTAL; ws.cell(row=tot_row, column=2).border = BORDER
ws.cell(row=tot_row, column=4, value=f"=SUM(D{hdr_r+1}:D{tot_row-1})"); style_total(ws.cell(row=tot_row, column=4)); ws.cell(row=tot_row, column=4).number_format = "0"
ws.cell(row=tot_row, column=5, value=f"=SUM(E{hdr_r+1}:E{tot_row-1})"); style_total(ws.cell(row=tot_row, column=5)); ws.cell(row=tot_row, column=5).number_format = "$#,##0"

# Other costs + total
oc_row = tot_row + 1
ws.cell(row=oc_row, column=2, value="Other costs (cloud, LLM, R&D compute, fine-tune, sandbox)").font = F_LBL
ws.cell(row=oc_row, column=2).alignment = ALIGN_L; ws.cell(row=oc_row, column=2).border = BORDER
oc = ws.cell(row=oc_row, column=5, value=45000); style_input(oc); oc.number_format = "$#,##0"

# Direct cost
dc_row = oc_row + 1
ws.cell(row=dc_row, column=2, value="Direct cost — sum").font = F_TOT
ws.cell(row=dc_row, column=2).fill = FILL_TOTAL; ws.cell(row=dc_row, column=2).border = BORDER
dc = ws.cell(row=dc_row, column=5, value=f"=E{tot_row}+E{oc_row}"); style_total(dc); dc.number_format = "$#,##0"

# Contingency
cont_row = dc_row + 1
ws.cell(row=cont_row, column=2, value="Contingency (15% — bespoke risk)").font = F_LBL
ws.cell(row=cont_row, column=2).alignment = ALIGN_L; ws.cell(row=cont_row, column=2).border = BORDER
cc = ws.cell(row=cont_row, column=5, value=f"=E{dc_row}*0.15"); style_formula(cc); cc.number_format = "$#,##0"

# Fully loaded — this is the C18 referenced from Executive Summary
loaded_row = cont_row + 1
ws.cell(row=loaded_row, column=2, value="Total fully-loaded bespoke cost").font = F_TOT
ws.cell(row=loaded_row, column=2).fill = FILL_TOTAL; ws.cell(row=loaded_row, column=2).border = BORDER
# Put both summary refs into column C for executive-summary boxes (cost in C, price in E)
fl_cost = ws.cell(row=loaded_row, column=3, value=f"=E{dc_row}+E{cont_row}")
style_total(fl_cost); fl_cost.number_format = "$#,##0"
fl_cost.font = Font(name="Calibri", size=11, bold=True, color=NAVY)

# Price = cost / (1 - 50% margin)
ws.cell(row=loaded_row, column=4, value="@ 50% margin →").font = F_TXT
ws.cell(row=loaded_row, column=4).alignment = ALIGN_R
fl_price = ws.cell(row=loaded_row, column=5, value=f"=C{loaded_row}/(1-'1. Cover & Assumptions'!$C$33)")
style_total(fl_price); fl_price.number_format = "$#,##0"
fl_price.font = Font(name="Calibri", size=12, bold=True, color=NAVY)

# Indicative T&M alternative pricing
tm_row = loaded_row + 3
ws.cell(row=tm_row, column=2, value="Alternative — T&M model").font = F_H1
ws.cell(row=tm_row + 1, column=2, value="Blended bill rate (USD/day)").font = F_LBL
ws.cell(row=tm_row + 1, column=2).alignment = ALIGN_L; ws.cell(row=tm_row + 1, column=2).border = BORDER
br = ws.cell(row=tm_row + 1, column=5, value=1500); style_input(br); br.number_format = "$#,##0"
ws.cell(row=tm_row + 1, column=7, value="Blended bill rate across the 15-person team. Editable.").font = F_NOTE

ws.cell(row=tm_row + 2, column=2, value="Estimated T&M billing for reference build").font = F_LBL
ws.cell(row=tm_row + 2, column=2).alignment = ALIGN_L; ws.cell(row=tm_row + 2, column=2).border = BORDER
tm_total = ws.cell(row=tm_row + 2, column=5, value=f"=D{tot_row}*'1. Cover & Assumptions'!$C$24*E{tm_row+1}")
style_formula(tm_total); tm_total.number_format = "$#,##0"
ws.cell(row=tm_row + 2, column=7, value="Total p-weeks × 5 days × blended bill rate. Compare to fixed-bid above.").font = F_NOTE


# =============================================================================
# SHEET 8 — Pricing Summary
# =============================================================================
ws = wb.create_sheet("8. Pricing Summary")
set_widths(ws, {"A": 3, "B": 46, "C": 18, "D": 18, "E": 18, "F": 60})

ws["B2"] = "Pricing Summary — Cost → Margin → Price to e&"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: Each tier's internal cost grossed up to a price using the target margin from Sheet 1. Price to e& = cost / (1 - margin)."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

hdr_r = 5
for i, h in enumerate(["Component", "Total cost", "Margin (USD)", "Price to e&", "Notes"]):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR_V; c.alignment = ALIGN_C; c.border = BORDER

# Cell references — Sheet 3: total p-weeks row 22, personnel cost row 25, other 26, direct 27, contingency 28, fully loaded 29.
# Sheet 4: Total cost per agent row 27 (col J = avg).
# Sheet 5: Total cost per project row 27 (col I = avg).
# Sheet 6: Direct subtotal row 15, contingency row 16, fully loaded row 17.
# Sheet 7: Total fully loaded bespoke cost row 26 (col C = cost, col E = price).
# Cover & Assumptions margin rows: Wave E0 = 29, Tier1 = 30, Tier2 = 31, Tier3 = 32, Custom = 33, Managed = 34.
#
# Pricing Summary layout (final):
#   Row 8  : "A. Wave E0" header
#   Row 9  : Wave E0 cost / margin / price       → Exec Summary refs C9, E9
#   Row 12 : "B. Tier 1" header
#   Row 13 : column headers
#   Row 14-19 : 6 Tier 1 agents
#   Row 20 : Tier 1 average                      → Exec Summary refs C20, E20
#   Row 23 : "C. Tier 2" header
#   Row 24 : column headers
#   Row 25-29 : 5 Tier 2 solutions (incl. AI Contact Centre)
#   Row 30 : Tier 2 average                      → Exec Summary refs C30, E30
#   Row 33 : "D. Tier 3" header
#   Row 34 : column headers
#   Row 36 : Tier 3 typical                      → Exec Summary refs C36, E36
#   Row 39 : "E. Managed Service" header
#   Row 40 : column headers
#   Row 41 : Managed service typical             → Exec Summary refs C41, E41

S3_LOADED = "'3. Wave E0 - Platform'!H29"
S6_LOADED = "'6. Tier 3 - Sovereign OnPrem'!C17"

# A. Wave E0
ws.cell(row=8, column=2, value="A.  Wave E0 — Enterprise Platform Extensions").font = F_H1

r = 9
ws.cell(row=r, column=2, value="Wave E0 — one-off platform build").font = F_LBL
ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
cc = ws.cell(row=r, column=3, value=f"={S3_LOADED}"); style_formula(cc); cc.number_format = "$#,##0"
mc = ws.cell(row=r, column=4, value=f"=C{r}/(1-'1. Cover & Assumptions'!$C$29)-C{r}"); style_formula(mc); mc.number_format = "$#,##0"
pc = ws.cell(row=r, column=5, value=f"=C{r}+D{r}"); style_total(pc); pc.number_format = "$#,##0"
pc.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
ws.cell(row=r, column=6, value="One-off charged to first 1-2 enterprise deals or amortised across the first 4-5 deals (commercial choice).").font = F_NOTE
ws.cell(row=r, column=6).border = BORDER

# B. Tier 1
ws.cell(row=12, column=2, value="B.  Tier 1 — Adapted SMB Agents (per agent)").font = F_H1
hdr_t1 = 13
for i, h in enumerate(["Agent (Tier 1)", "Cost", "Margin", "Price to e&", "Notes"]):
    c = ws.cell(row=hdr_t1, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

t1_agents = ["Customer Agent", "Sales Agent", "Comms Hub", "Finance Agent", "Ops Agent", "People Agent"]
for i, agent in enumerate(t1_agents):
    r = hdr_t1 + 1 + i  # rows 14..19
    ws.cell(row=r, column=2, value=agent).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    col_letter = get_column_letter(4 + i)
    cc = ws.cell(row=r, column=3, value=f"='4. Tier 1 - Adapted Agents'!{col_letter}27")
    style_formula(cc); cc.number_format = "$#,##0"
    mc = ws.cell(row=r, column=4, value=f"=C{r}/(1-'1. Cover & Assumptions'!$C$30)-C{r}"); style_formula(mc); mc.number_format = "$#,##0"
    pc = ws.cell(row=r, column=5, value=f"=C{r}+D{r}"); style_total(pc); pc.number_format = "$#,##0"
    ws.cell(row=r, column=6, value=f"Tier 1 enterprise edition of the SMB {agent}.").font = F_NOTE
    ws.cell(row=r, column=6).border = BORDER

# Tier 1 average — row 20 (used by Exec Summary and Sheet 9 / 11)
r = 20
ws.cell(row=r, column=2, value="↑ Tier 1 average per agent").font = F_TOT
ws.cell(row=r, column=2).fill = FILL_TOTAL; ws.cell(row=r, column=2).border = BORDER
ac = ws.cell(row=r, column=3, value=f"=AVERAGE(C{hdr_t1+1}:C{r-1})"); style_total(ac); ac.number_format = "$#,##0"
am = ws.cell(row=r, column=4, value=f"=AVERAGE(D{hdr_t1+1}:D{r-1})"); style_total(am); am.number_format = "$#,##0"
ap = ws.cell(row=r, column=5, value=f"=AVERAGE(E{hdr_t1+1}:E{r-1})"); style_total(ap); ap.number_format = "$#,##0"
ap.font = Font(name="Calibri", size=11, bold=True, color=NAVY)

# C. Tier 2
ws.cell(row=23, column=2, value="C.  Tier 2 — Custom AI Solutions (per project)").font = F_H1
hdr_t2 = 24
for i, h in enumerate(["Solution (Tier 2)", "Cost", "Margin", "Price to e&", "Notes"]):
    c = ws.cell(row=hdr_t2, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

t2_solutions = ["AI Contact Centre Software", "Document Intelligence", "Approval Orchestration", "Predictive Maintenance", "Security & Compliance"]
for i, sol in enumerate(t2_solutions):
    r = hdr_t2 + 1 + i  # rows 25..29
    ws.cell(row=r, column=2, value=sol).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    col_letter = get_column_letter(4 + i)
    cc = ws.cell(row=r, column=3, value=f"='5. Tier 2 - Custom AI'!{col_letter}27")
    style_formula(cc); cc.number_format = "$#,##0"
    mc = ws.cell(row=r, column=4, value=f"=C{r}/(1-'1. Cover & Assumptions'!$C$31)-C{r}"); style_formula(mc); mc.number_format = "$#,##0"
    pc = ws.cell(row=r, column=5, value=f"=C{r}+D{r}"); style_total(pc); pc.number_format = "$#,##0"
    ws.cell(row=r, column=6, value=f"Tier 2 — fits in JSX framework $300K-$600K range.").font = F_NOTE
    ws.cell(row=r, column=6).border = BORDER

# Tier 2 average — row 30
r = 30
ws.cell(row=r, column=2, value="↑ Tier 2 average per project").font = F_TOT
ws.cell(row=r, column=2).fill = FILL_TOTAL; ws.cell(row=r, column=2).border = BORDER
ac = ws.cell(row=r, column=3, value=f"=AVERAGE(C{hdr_t2+1}:C{r-1})"); style_total(ac); ac.number_format = "$#,##0"
am = ws.cell(row=r, column=4, value=f"=AVERAGE(D{hdr_t2+1}:D{r-1})"); style_total(am); am.number_format = "$#,##0"
ap = ws.cell(row=r, column=5, value=f"=AVERAGE(E{hdr_t2+1}:E{r-1})"); style_total(ap); ap.number_format = "$#,##0"
ap.font = Font(name="Calibri", size=11, bold=True, color=NAVY)

# D. Tier 3
ws.cell(row=33, column=2, value="D.  Tier 3 — Sovereign / On-Prem (per project)").font = F_H1
hdr_t3 = 34
for i, h in enumerate(["Component", "Cost", "Margin", "Price to e&", "Notes"]):
    c = ws.cell(row=hdr_t3, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Tier 3 typical project — row 36
r = 36
ws.cell(row=r, column=2, value="Tier 3 — typical sovereign / on-prem build").font = F_LBL
ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
cc = ws.cell(row=r, column=3, value=f"={S6_LOADED}"); style_formula(cc); cc.number_format = "$#,##0"
mc = ws.cell(row=r, column=4, value=f"=C{r}/(1-'1. Cover & Assumptions'!$C$32)-C{r}"); style_formula(mc); mc.number_format = "$#,##0"
pc = ws.cell(row=r, column=5, value=f"=C{r}+D{r}"); style_total(pc); pc.number_format = "$#,##0"
pc.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
ws.cell(row=r, column=6, value="Tier 3 typically prices in the $600K-$1M JSX band. Hardware BoM sold separately by OEM. See Sheet 6 for the cost stack.").font = F_NOTE
ws.cell(row=r, column=6).border = BORDER

# E. Managed Service
ws.cell(row=39, column=2, value="E.  Managed Service (per client per month)").font = F_H1
hdr_ms = 40
for i, h in enumerate(["Component", "Cost / month", "Margin", "Price to e& / month", "Notes"]):
    c = ws.cell(row=hdr_ms, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Managed service typical — row 41
# Cost estimate: ~0.2 PM + ~0.3 AI Eng + ~0.1 DevOps + ~0.2 CS dedicated per client
# Roughly: 0.2*$1000*20 + 0.3*$900*20 + 0.1*$950*20 + 0.2*$600*20 = $4000+$5400+$1900+$2400 = ~$13.7K/mo (high)
# We use $5400/mo (mostly L3 escalations + tooling) which targets $12K ARPU at 55% margin.
r = 41
ws.cell(row=r, column=2, value="Managed service — typical").font = F_LBL
ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
mc = ws.cell(row=r, column=3, value=5400); style_input(mc); mc.number_format = "$#,##0"
mm = ws.cell(row=r, column=4, value=f"=C{r}/(1-'1. Cover & Assumptions'!$C$34)-C{r}"); style_formula(mm); mm.number_format = "$#,##0"
mp = ws.cell(row=r, column=5, value=f"=C{r}+D{r}"); style_total(mp); mp.number_format = "$#,##0"
mp.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
ws.cell(row=r, column=6, value="Cross-check: assumption ARPU = '1. Cover & Assumptions'!$C$50 ($12K/mo). Editable cost cell.").font = F_NOTE
ws.cell(row=r, column=6).border = BORDER

# F. JSX Reality Check — internal price vs JSX commercial framework target
ws.cell(row=44, column=2, value="F.  Reality check — internal price vs JSX commercial framework").font = F_H1
hdr_rc = 45
for i, h in enumerate(["Component", "Internal price", "JSX low", "JSX high", "Notes"]):
    c = ws.cell(row=hdr_rc, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

reality_check = [
    ("Tier 1 — average per agent",        "=E20", 30000, 80000,
     "Internal model is well above JSX top end — JSX assumed near-trivial config; reality requires real ERP integration. Reduce p-weeks if rolling out the same agent to a customer's 2nd/3rd seat (ERP work amortised)."),
    ("Tier 2 — average per project",       "=E30", 300000, 600000,
     "Mid range fits, but the AI Contact Centre solution lands above the JSX top end. Negotiable — see Sheet 5."),
    ("Tier 3 — typical per project",       "=E36", 600000, 1000000,
     "Internal model just above JSX top end. Sovereign deals routinely exceed $1M; ok to set Tier 3 ceiling at $1.2-1.5M for premium gov/bank flagships."),
    ("Custom bespoke",                      "='7. Custom Bespoke'!E26", 400000, 800000,
     "JSX did not publish a custom bespoke price band. Use Sheet 7 fixed-bid or T&M model per opportunity."),
    ("Managed service per client / month", "=E41", 5000, 25000,
     "$12K/mo lands in the JSX $5K-$25K band ✓"),
]
for i, (label, internal, low, high, note) in enumerate(reality_check):
    r = hdr_rc + 1 + i
    ws.cell(row=r, column=2, value=label).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    ic = ws.cell(row=r, column=3, value=internal); style_total(ic); ic.number_format = "$#,##0"
    ic.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    lc = ws.cell(row=r, column=4, value=low); style_input(lc); lc.number_format = "$#,##0"
    hc = ws.cell(row=r, column=5, value=high); style_input(hc); hc.number_format = "$#,##0"
    nc = ws.cell(row=r, column=6, value=note); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER

# Tier 1 amortised — second & third agent for same customer
ws.cell(row=52, column=2, value="G.  Tier 1 amortised — 2nd / 3rd agent for the same enterprise customer").font = F_H1
hdr_am = 53
for i, h in enumerate(["Scenario", "Cost (USD)", "Margin", "Price to e&", "Notes"]):
    c = ws.cell(row=hdr_am, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

amort_rows = [
    ("Tier 1 — 1st agent for new customer (full)", "=C20", "Full ERP/SSO/audit work."),
    ("Tier 1 — 2nd agent for same customer (~60%)", "=C20*0.60", "ERP connector reused; only agent-specific config + UAT."),
    ("Tier 1 — 3rd+ agent for same customer (~50%)", "=C20*0.50", "Most integration plumbing already in place; mostly prompt tuning + UAT."),
]
for i, (label, val, note) in enumerate(amort_rows):
    r = hdr_am + 1 + i
    ws.cell(row=r, column=2, value=label).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    cc = ws.cell(row=r, column=3, value=val); style_formula(cc); cc.number_format = "$#,##0"
    mc = ws.cell(row=r, column=4, value=f"=C{r}/(1-'1. Cover & Assumptions'!$C$30)-C{r}"); style_formula(mc); mc.number_format = "$#,##0"
    pc = ws.cell(row=r, column=5, value=f"=C{r}+D{r}"); style_total(pc); pc.number_format = "$#,##0"
    pc.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    nc = ws.cell(row=r, column=6, value=note); nc.font = F_NOTE; nc.alignment = ALIGN_L; nc.border = BORDER


# =============================================================================
# SHEET 9 — 3-Year Revenue Forecast
# =============================================================================
ws = wb.create_sheet("9. Revenue Forecast")
set_widths(ws, {"A": 3, "B": 46, "C": 18, "D": 18, "E": 18, "F": 18, "G": 50})

ws["B2"] = "3-Year Revenue Forecast — Pipeline × Tier Mix → AIdeology vs e&"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: Number of deals per year × tier mix × price per tier = gross revenue. Apply 60/40 AIdeology/e& split. Add managed service ramp."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

# Deal volumes
ws["B5"] = "A. Deal volumes (mid scenario)"
ws["B5"].font = F_H1

hdr_r = 6
for i, h in enumerate(["Metric", "Year 1", "Year 2", "Year 3", "3-Year Total", "Notes"]):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Total deals
r = 9  # row 9 for exec summary
ws.cell(row=r, column=2, value="# Total deals").font = F_LBL; ws.cell(row=r, column=2).border = BORDER
ws.cell(row=r, column=3, value="='1. Cover & Assumptions'!$C$42"); style_formula(ws.cell(row=r, column=3)); ws.cell(row=r, column=3).number_format = "#,##0"
ws.cell(row=r, column=4, value="='1. Cover & Assumptions'!$C$43"); style_formula(ws.cell(row=r, column=4)); ws.cell(row=r, column=4).number_format = "#,##0"
ws.cell(row=r, column=5, value="='1. Cover & Assumptions'!$C$44"); style_formula(ws.cell(row=r, column=5)); ws.cell(row=r, column=5).number_format = "#,##0"
ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})"); style_total(ws.cell(row=r, column=6)); ws.cell(row=r, column=6).number_format = "#,##0"

# Deal mix breakdown
mix_breakdown = [
    ("# Tier 1 deals (= Total × Tier1 mix %)",       "C9*'1. Cover & Assumptions'!$C$45", "D9*'1. Cover & Assumptions'!$C$45", "E9*'1. Cover & Assumptions'!$C$45"),
    ("# Tier 1 agents in pipeline (× avg agents/deal)", "C10*'1. Cover & Assumptions'!$C$49",   "D10*'1. Cover & Assumptions'!$C$49",   "E10*'1. Cover & Assumptions'!$C$49"),
    ("# Tier 2 deals",                                "C9*'1. Cover & Assumptions'!$C$46",  "D9*'1. Cover & Assumptions'!$C$46",  "E9*'1. Cover & Assumptions'!$C$46"),
    ("# Tier 3 deals",                                "C9*'1. Cover & Assumptions'!$C$47",  "D9*'1. Cover & Assumptions'!$C$47",  "E9*'1. Cover & Assumptions'!$C$47"),
    ("# Custom bespoke deals",                        "C9*'1. Cover & Assumptions'!$C$48",  "D9*'1. Cover & Assumptions'!$C$48",  "E9*'1. Cover & Assumptions'!$C$48"),
]
for i, (label, y1, y2, y3) in enumerate(mix_breakdown):
    r = hdr_r + 4 + i  # 10..14
    ws.cell(row=r, column=2, value=label).font = F_LBL; ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=3, value=f"={y1}"); style_formula(ws.cell(row=r, column=3)); ws.cell(row=r, column=3).number_format = "0.0"
    ws.cell(row=r, column=4, value=f"={y2}"); style_formula(ws.cell(row=r, column=4)); ws.cell(row=r, column=4).number_format = "0.0"
    ws.cell(row=r, column=5, value=f"={y3}"); style_formula(ws.cell(row=r, column=5)); ws.cell(row=r, column=5).number_format = "0.0"
    ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})"); style_total(ws.cell(row=r, column=6)); ws.cell(row=r, column=6).number_format = "0.0"

# Revenue section
ws.cell(row=17, column=2, value="B. Gross revenue to e& (price × volume)").font = F_H1
hdr_rev = 18
for i, h in enumerate(["Stream", "Year 1", "Year 2", "Year 3", "3-Year Total", "Notes"]):
    c = ws.cell(row=hdr_rev, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Total project revenue (row 19 for exec summary)
# = Tier1_agents × T1 price + Tier2 × T2 price + Tier3 × T3 price + Custom × Custom price
T1_PRICE = "'8. Pricing Summary'!$E$20"
T2_PRICE = "'8. Pricing Summary'!$E$30"
T3_PRICE = "'8. Pricing Summary'!$E$36"
CUSTOM_PRICE = "'7. Custom Bespoke'!$E$26"  # Custom bespoke price (row 26 = total fully-loaded; col E = price after 50% margin gross-up)

r = 19  # row 19 for exec summary
ws.cell(row=r, column=2, value="Gross project revenue (build fees)").font = F_TOT
ws.cell(row=r, column=2).fill = FILL_TOTAL; ws.cell(row=r, column=2).border = BORDER
for j in range(3):
    col = get_column_letter(3 + j)
    formula = f"={col}11*{T1_PRICE}+{col}12*{T2_PRICE}+{col}13*{T3_PRICE}+{col}14*{CUSTOM_PRICE}"
    cell = ws.cell(row=r, column=3 + j, value=formula)
    style_total(cell); cell.number_format = "$#,##0"
tot_cell = ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})")
style_total(tot_cell); tot_cell.number_format = "$#,##0"
tot_cell.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
ws.cell(row=r, column=7, value="Tier1 agents × T1 + Tier2 × T2 + Tier3 × T3 + Custom × Custom").font = F_NOTE

# Managed service revenue ramp
ws.cell(row=22, column=2, value="C. Managed service revenue (recurring)").font = F_H1
ws.cell(row=23, column=2, value="Cumulative managed-service clients").font = F_LBL; ws.cell(row=23, column=2).border = BORDER
# Cumulative clients: assume each year's deals → uptake % of clients sign managed service, with a 1-month lag (so partial year for new ones)
# Simpler model: managed clients in Y_t = sum(deals 1..t) × uptake%. Average over the year = 0.5 × beginning + 0.5 × end (roughly 0.5 of new)
ws.cell(row=23, column=3, value="=C9*'1. Cover & Assumptions'!$C$51*0.5"); style_formula(ws.cell(row=23, column=3)); ws.cell(row=23, column=3).number_format = "0.0"
ws.cell(row=23, column=4, value="=C9*'1. Cover & Assumptions'!$C$51 + D9*'1. Cover & Assumptions'!$C$51*0.5"); style_formula(ws.cell(row=23, column=4)); ws.cell(row=23, column=4).number_format = "0.0"
ws.cell(row=23, column=5, value="=(C9+D9)*'1. Cover & Assumptions'!$C$51 + E9*'1. Cover & Assumptions'!$C$51*0.5"); style_formula(ws.cell(row=23, column=5)); ws.cell(row=23, column=5).number_format = "0.0"
ws.cell(row=23, column=7, value="Avg paying clients during the year (half-year ramp for new ones).").font = F_NOTE

# Row 24 for exec summary — managed revenue
r = 24
ws.cell(row=r, column=2, value="Managed service revenue").font = F_TOT
ws.cell(row=r, column=2).fill = FILL_TOTAL; ws.cell(row=r, column=2).border = BORDER
for j in range(3):
    col = get_column_letter(3 + j)
    cell = ws.cell(row=r, column=3 + j, value=f"={col}23*'1. Cover & Assumptions'!$C$50*12")
    style_total(cell); cell.number_format = "$#,##0"
tot_cell = ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})")
style_total(tot_cell); tot_cell.number_format = "$#,##0"
tot_cell.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
ws.cell(row=r, column=7, value="Avg clients × ARPU × 12 months (managed service ARPU $12K/mo).").font = F_NOTE

# Total revenue
ws.cell(row=27, column=2, value="D. Total revenue and split").font = F_H1
hdr_split = 28
for i, h in enumerate(["Item", "Year 1", "Year 2", "Year 3", "3-Year Total", "Notes"]):
    c = ws.cell(row=hdr_split, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# Gross combined
r = 29
ws.cell(row=r, column=2, value="Gross combined revenue (project + managed)").font = F_TOT
ws.cell(row=r, column=2).fill = FILL_TOTAL; ws.cell(row=r, column=2).border = BORDER
for j in range(3):
    col = get_column_letter(3 + j)
    cell = ws.cell(row=r, column=3 + j, value=f"={col}19+{col}24")
    style_total(cell); cell.number_format = "$#,##0"
tc = ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})"); style_total(tc); tc.number_format = "$#,##0"

# AIdeology share (row 30 for exec summary)
r = 30
ws.cell(row=r, column=2, value="AIdeology share (60% project + 60% managed)").font = F_TOT
ws.cell(row=r, column=2).fill = FILL_GREEN_LT; ws.cell(row=r, column=2).border = BORDER
for j in range(3):
    col = get_column_letter(3 + j)
    cell = ws.cell(row=r, column=3 + j, value=f"={col}29*'1. Cover & Assumptions'!$C$39")
    style_total(cell); cell.number_format = "$#,##0"
    cell.fill = FILL_GREEN_LT
    cell.font = Font(name="Calibri", size=12, bold=True, color=NAVY)
tc = ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})"); style_total(tc); tc.number_format = "$#,##0"
tc.font = Font(name="Calibri", size=14, bold=True, color=NAVY); tc.fill = FILL_GREEN_LT

# e& share (row 31 for exec summary)
r = 31
ws.cell(row=r, column=2, value="e& share (40% project + 40% managed; hosting separate)").font = F_TOT
ws.cell(row=r, column=2).fill = FILL_VIOLET; ws.cell(row=r, column=2).border = BORDER
for j in range(3):
    col = get_column_letter(3 + j)
    cell = ws.cell(row=r, column=3 + j, value=f"={col}29*'1. Cover & Assumptions'!$C$40")
    style_total(cell); cell.number_format = "$#,##0"
    cell.fill = FILL_VIOLET
tc = ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})"); style_total(tc); tc.number_format = "$#,##0"
tc.fill = FILL_VIOLET


# =============================================================================
# SHEET 10 — HPC RA Library
# =============================================================================
ws = wb.create_sheet("10. HPC RA Library")
set_widths(ws, {"A": 3, "B": 22, "C": 36, "D": 28, "E": 26, "F": 18, "G": 14, "H": 22, "I": 60})

ws["B2"] = "HPC Reference Architecture Library — pre-prepared BoMs for vendor agility"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: AIdeology has prepared NVIDIA-Certified reference architectures for every major OEM. When a Tier 3 deal lands, we pick the matching RA and produce a customer-specific BoM in days, not weeks."
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

# RA family overview
ws["B5"] = "A. NVIDIA Enterprise RA families"
ws["B5"].font = F_H1

hdr_r = 6
for i, h in enumerate(["Family", "Pattern (C-G-N-B)", "GPU", "Default size (4-node SU)", "Max size", "Best fit"]):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

ra_families = [
    ("Enterprise RA · RTX PRO", "2-8-5-200", "RTX PRO 6000 Blackwell SE", "4 nodes / 32 GPUs", "32 nodes / 256 GPUs", "Mid-density inference, agentic AI, fine-tuning at scale, dept-level GPUaaS"),
    ("HGX AI Factory", "2-8-9-400", "HGX H200 / B200", "4 nodes / 32 GPUs", "128 nodes / 1,024 GPUs", "Workhorse training & fine-tuning, sovereign GPUaaS, regulated training jobs"),
    ("HGX B300 Single Plane", "2-8-9-800", "HGX B300", "4 nodes / 32 GPUs", "32 nodes / 256 GPUs", "Frontier model training, advanced inference, AI factory build-out"),
]
for i, (fam, pattern, gpu, default, mx, fit) in enumerate(ra_families):
    r = hdr_r + 1 + i
    cells = [fam, pattern, gpu, default, mx, fit]
    for j, v in enumerate(cells):
        c = ws.cell(row=r, column=2 + j, value=v)
        c.font = F_TXT; c.alignment = ALIGN_L; c.border = BORDER

# OEM matrix
oem_lbl_row = hdr_r + 1 + len(ra_families) + 2
ws.cell(row=oem_lbl_row, column=2, value="B. OEM-validated solutions (pre-prepared BoMs in hand)").font = F_H1
hdr_o = oem_lbl_row + 1
for i, h in enumerate(["Vendor", "Solution", "Server", "GPU", "Pattern", "Size range", "Endorsements", "Notes"]):
    c = ws.cell(row=hdr_o, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

oems = [
    ("Cisco", "Nexus Hyperfabric AI ERA", "UCS C885A Rack Server", "HGX H200", "2-8-9-400 → 2-8-10-400", "4 – 128", "Infra · Spectrum-X · Networking", "Deep telco-network alignment with e&; preferred for sovereign GPUaaS."),
    ("Cisco", "Nexus 9000 ERA", "UCS C885A Rack Server", "HGX H200", "2-8-9-400", "4 – 128", "Infra · Spectrum-X · Networking", "Same UCS server, alternate fabric for existing Nexus 9000 customers."),
    ("Cisco", "AI POD Infrastructure", "UCS C885A M8", "HGX H200", "2-8-9-400", "4 – 16", "Infra · Spectrum-X", "Smaller POD; ideal for first-deployment on-prem PoC or single-customer enterprise."),
    ("Dell Technologies", "Dell AI Factory with NVIDIA", "PowerEdge XE7745", "RTX PRO 6000 Blackwell SE / H200 NVL", "2-8-5-200", "4, 16", "Infra · Spectrum-X · Networking", "Strongest agentic AI reference. Mid-density inference workhorse."),
    ("Dell Technologies", "Dell AI Factory with NVIDIA", "PowerEdge XE7740", "RTX PRO 6000 Blackwell SE / H200 NVL", "2-8-5-200", "4, 16", "Infra · Spectrum-X · Networking", "Slightly different chassis; inference-leaning."),
    ("Dell Technologies", "Dell AI Factory with NVIDIA", "PowerEdge XE9680", "HGX H200", "2-8-9-400", "4, 12", "Infra · Spectrum-X · Networking", "HGX-class training. Up to 12 SUs / 96 GPUs."),
    ("Dell Federal", "Dell AI Factory for Government", "PowerEdge XE7740 / XE7745", "RTX PRO 6000 Blackwell SE", "2-8-5-200", "4 – 32", "Infra · Spectrum-X", "Government-validated SKU; FedRAMP-aligned procurement."),
    ("Lenovo", "Lenovo Hybrid AI 289", "ThinkSystem SR680a V3", "HGX B200 / H200", "2-8-9-400", "4 – 32", "Infra · Spectrum-X · Networking", "Strong B200 path; co-engineered with NVIDIA on hybrid AI."),
    ("Supermicro", "AI Factory · HGX", "SYS-A22GA-NBRT-G1", "HGX B200", "2-8-9-400", "4 – 32", "Infra · Spectrum-X", "Aggressive pricing on HGX class; favoured for fine-tune-heavy tenants."),
    ("Supermicro", "AI Factory · RTX PRO 6000", "SYS-522GA-NRT / SYS-422GL-NR / AS-5126GS-TNRT2", "RTX PRO 6000 Blackwell SE", "2-8-5-200", "4 – 32", "Infra · Spectrum-X", "Mid-density inference; lowest TCO of RTX PRO RAs."),
    ("Supermicro", "AI Factory · HGX B300 Single Plane", "SYS-822GS-NB3RT", "HGX B300 Single Plane", "2-8-9-800", "4, 8, 32", "Infra · Spectrum-X", "Frontier B300 — best for any new-flagship training opportunity."),
]
for i, oem in enumerate(oems):
    r = hdr_o + 1 + i
    for j, v in enumerate(oem):
        c = ws.cell(row=r, column=2 + j, value=v)
        c.font = F_TXT; c.alignment = ALIGN_L; c.border = BORDER

# Indicative BoM costs
bom_lbl_row = hdr_o + 1 + len(oems) + 2
ws.cell(row=bom_lbl_row, column=2, value="C. Indicative BoM list-price ranges (vendor MSRP — e& negotiating discount typically 20-35%)").font = F_H1
hdr_b = bom_lbl_row + 1
for i, h in enumerate(["BoM size (GPUs)", "Pattern", "List-price range (USD)", "Indicative compute & storage", "AIdeology BoM coordination fee"]):
    c = ws.cell(row=hdr_b, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

bom_sizes = [
    ("32 GPUs · 4 nodes (1 SU)", "2-8-5-200 (RTX PRO)", "$1.5M – $2.5M", "Mid-density inference + small fine-tune. ~250 TB AI storage.", "Included in Tier 3 Sheet 6"),
    ("32 GPUs · 4 nodes (1 SU)", "2-8-9-400 (HGX H200)", "$2.5M – $3.5M", "Training + fine-tune workhorse. ~500 TB AI storage.", "Included in Tier 3 Sheet 6"),
    ("128 GPUs · 16 nodes (4 SUs)", "2-8-5-200 (RTX PRO)", "$5M – $8M", "Departmental GPUaaS; multiple tenants.", "+$25K (allowance — Sheet 6)"),
    ("256 GPUs · 32 nodes (8 SUs / Standard Pod)", "2-8-9-400 (HGX H200/B200)", "$15M – $22M", "Standard fully-tested pod. Sovereign GPUaaS or flagship enterprise.", "+$50K (uplift)"),
    ("1,024 GPUs · 128 nodes (32 SUs / Max Pod)", "2-8-9-400 (HGX H200/B200)", "$60M – $90M", "Frontier-scale, multi-pod. Government / consortium build-out.", "+$150K (uplift)"),
]
for i, bom in enumerate(bom_sizes):
    r = hdr_b + 1 + i
    for j, v in enumerate(bom):
        c = ws.cell(row=r, column=2 + j, value=v)
        c.font = F_TXT; c.alignment = ALIGN_L; c.border = BORDER

note_row = hdr_b + 1 + len(bom_sizes) + 2
ws.cell(row=note_row, column=2, value="Reading guidance").font = F_H1
ws.cell(row=note_row + 1, column=2, value="• AIdeology does not sell hardware — the OEM or e& takes that line. Our role is BoM coordination + integration, costed in Sheet 6.").font = F_NOTE
ws.cell(row=note_row + 2, column=2, value="• For deals > 256 GPUs, AIdeology charges a BoM coordination uplift (column 6 above) on top of the standard Tier 3 cost stack.").font = F_NOTE
ws.cell(row=note_row + 3, column=2, value="• Source: NVIDIA Enterprise Reference Architecture white paper — docs.nvidia.com/enterprise-reference-architectures.").font = F_NOTE


# =============================================================================
# SHEET 11 — Sensitivity
# =============================================================================
ws = wb.create_sheet("11. Sensitivity")
set_widths(ws, {"A": 3, "B": 38, "C": 16, "D": 16, "E": 16, "F": 18, "G": 50})

ws["B2"] = "Sensitivity — Pipeline volume × tier mix × revenue split"
ws["B2"].font = F_TITLE
ws["B3"] = "IN PLAIN ENGLISH: How does AIdeology's 3-year revenue change under different pipeline assumptions and revenue splits?"
ws["B3"].font = Font(name="Calibri", size=11, color="FF2E75B6", italic=True)

# A. Pipeline scenarios
ws["B5"] = "A. Pipeline scenarios — 3-year deal volume"
ws["B5"].font = F_H1

hdr_r = 6
for i, h in enumerate(["Scenario", "Year 1 deals", "Year 2 deals", "Year 3 deals", "3-Year Total", "Notes"]):
    c = ws.cell(row=hdr_r, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

scenarios = [
    ("Low — slow ramp (3, 8, 15)",  3, 8, 15,  "Conservative: 3-8-15 deals. JSX low end."),
    ("Mid — baseline (4, 11, 18)",   4, 11, 18, "Baseline used in Sheet 9. JSX mid."),
    ("High — strong adoption (8, 15, 25)", 8, 15, 25, "Aggressive: 8-15-25 deals. JSX high end."),
]
for i, (label, y1, y2, y3, note) in enumerate(scenarios):
    r = hdr_r + 1 + i
    ws.cell(row=r, column=2, value=label).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    for j, v in enumerate([y1, y2, y3]):
        cell = ws.cell(row=r, column=3 + j, value=v)
        style_input(cell); cell.number_format = "#,##0"
    tot = ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})"); style_total(tot); tot.number_format = "#,##0"
    ws.cell(row=r, column=7, value=note).font = F_NOTE; ws.cell(row=r, column=7).alignment = ALIGN_L; ws.cell(row=r, column=7).border = BORDER

# B. Avg revenue per deal (mix-weighted)
arpd_row = hdr_r + 1 + len(scenarios) + 2
ws.cell(row=arpd_row, column=2, value="B. Average revenue per deal (mix-weighted)").font = F_H1
ws.cell(row=arpd_row + 1, column=2, value="Tier 1 avg agents per deal × T1 price + Tier 2 + Tier 3 + Custom").font = F_LBL
ws.cell(row=arpd_row + 1, column=2).alignment = ALIGN_L; ws.cell(row=arpd_row + 1, column=2).border = BORDER
formula = ("="
           "'1. Cover & Assumptions'!$C$45 * '1. Cover & Assumptions'!$C$49 * '8. Pricing Summary'!$E$20 + "
           "'1. Cover & Assumptions'!$C$46 * '8. Pricing Summary'!$E$30 + "
           "'1. Cover & Assumptions'!$C$47 * '8. Pricing Summary'!$E$36 + "
           "'1. Cover & Assumptions'!$C$48 * '7. Custom Bespoke'!$E$26")
arpd_cell = ws.cell(row=arpd_row + 1, column=3, value=formula)
style_total(arpd_cell); arpd_cell.number_format = "$#,##0"
arpd_cell.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
ws.cell(row=arpd_row + 1, column=7, value="Single number = expected gross revenue per deal at current Tier mix.").font = F_NOTE

# C. AIdeology 3-year revenue under scenario × split
mt_row = arpd_row + 4
ws.cell(row=mt_row, column=2, value="C. AIdeology 3-year revenue (USD) — Scenario × Revenue split").font = F_H1
hdr_m = mt_row + 1
for i, h in enumerate(["Pipeline scenario", "50/50 split", "60/40 split (mid)", "70/30 split", "Notes"]):
    c = ws.cell(row=hdr_m, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

# For each scenario: total deals × avg deal revenue × split %, plus managed service contribution roughly proportional
# Simplify: total project revenue = total deals × avg revenue per deal. Managed service ramped, ~30% additional.
ms_factor = 0.30  # managed service is ~30% additional revenue across 3 years (proportional)
for i, (label, y1, y2, y3, _) in enumerate(scenarios):
    r = hdr_m + 1 + i
    ws.cell(row=r, column=2, value=label).font = F_LBL
    ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    sc_row = hdr_r + 1 + i
    total_deals_ref = f"F{sc_row}"  # 3-year total deals
    project_rev = f"{total_deals_ref}*$C${arpd_row+1}"
    combined_rev = f"({project_rev})*(1+{ms_factor})"
    for j, split in enumerate([0.50, 0.60, 0.70]):
        cell = ws.cell(row=r, column=3 + j, value=f"={combined_rev}*{split}")
        style_total(cell); cell.number_format = "$#,##0"
        if abs(split - 0.60) < 0.01 and "Mid" in label:
            cell.fill = FILL_GREEN_LT
            cell.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    ws.cell(row=r, column=7, value="Total deals × avg revenue × (1 + 30% managed) × split %").font = F_NOTE
    ws.cell(row=r, column=7).border = BORDER

# D. Tier mix sensitivity
ws.cell(row=hdr_m + 5, column=2, value="D. Tier mix sensitivity (per-deal revenue change)").font = F_H1
hdr_tm = hdr_m + 6
for i, h in enumerate(["Mix scenario", "T1 %", "T2 %", "T3 %", "Avg revenue / deal"]):
    c = ws.cell(row=hdr_tm, column=2 + i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER

mix_scenarios = [
    ("Tier 1 heavy (small deals)", 0.55, 0.30, 0.10, 0.05),
    ("Baseline", 0.35, 0.45, 0.15, 0.05),
    ("Tier 3 heavy (flagships)", 0.20, 0.40, 0.35, 0.05),
]
for i, (label, t1p, t2p, t3p, cp) in enumerate(mix_scenarios):
    r = hdr_tm + 1 + i
    ws.cell(row=r, column=2, value=label).font = F_LBL; ws.cell(row=r, column=2).alignment = ALIGN_L; ws.cell(row=r, column=2).border = BORDER
    t1c = ws.cell(row=r, column=3, value=t1p); style_input(t1c); t1c.number_format = "0%"
    t2c = ws.cell(row=r, column=4, value=t2p); style_input(t2c); t2c.number_format = "0%"
    t3c = ws.cell(row=r, column=5, value=t3p); style_input(t3c); t3c.number_format = "0%"
    arpd = ws.cell(row=r, column=6, value=(
        f"=C{r}*'1. Cover & Assumptions'!$C$49*'8. Pricing Summary'!$E$20 + "
        f"D{r}*'8. Pricing Summary'!$E$30 + "
        f"E{r}*'8. Pricing Summary'!$E$36 + "
        f"{cp}*'7. Custom Bespoke'!$E$26"
    ))
    style_total(arpd); arpd.number_format = "$#,##0"


# =============================================================================
# Save
# =============================================================================
import os
os.makedirs("Internal Analysis", exist_ok=True)
wb.save(OUT)
print(f"✓ Saved: {OUT}")
